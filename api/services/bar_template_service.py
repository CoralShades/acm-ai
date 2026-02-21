"""BAR Template Service — parse and manage BAR Excel templates.

Story: E5-S3 BAR Template Management
"""

from io import BytesIO
from typing import Optional

import openpyxl
from loguru import logger

from open_notebook.domain.bar_template import BARTemplate, BARTemplateColumn


def _col_letter(col_idx: int) -> str:
    """Convert 1-based column index to Excel column letter (A, B, ..., Z, AA, ...)."""
    result = ""
    while col_idx > 0:
        col_idx, remainder = divmod(col_idx - 1, 26)
        result = chr(65 + remainder) + result
    return result


# BAR required columns are A-AH (columns 1-34)
_REQUIRED_COLUMN_COUNT = 34


def parse_bar_template(file_bytes: bytes, filename: str) -> BARTemplate:
    """Parse a BAR Excel template (.xlsx/.xlsm) and extract column structure.

    Opens the workbook, finds the register sheet, reads the header row,
    and builds a BARTemplate with column metadata.
    """
    wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)

    # Find the register sheet
    ws = None
    for name in wb.sheetnames:
        if "register" in name.lower():
            ws = wb[name]
            break
    if ws is None:
        ws = wb.active

    if ws is None:
        raise ValueError("No worksheet found in the uploaded file")

    # Scan first 5 rows for the header row (first non-empty row)
    header_row = None
    header_row_idx = 0
    for row_idx in range(1, 6):
        row_values = []
        for cell in ws[row_idx]:
            row_values.append(cell.value)
        # Check if this row has enough non-None values to be a header
        non_empty = [v for v in row_values if v is not None]
        if len(non_empty) >= 5:  # At least 5 columns to be considered a header
            header_row = row_values
            header_row_idx = row_idx
            break

    if header_row is None:
        raise ValueError(
            "Could not find header row in the first 5 rows of the worksheet"
        )

    # Build columns from header
    columns = []
    raw_header = []
    for col_idx, value in enumerate(header_row, start=1):
        if value is None:
            continue
        col_name = str(value).strip()
        if not col_name:
            continue
        raw_header.append(col_name)
        columns.append(
            BARTemplateColumn(
                column_letter=_col_letter(col_idx),
                column_name=col_name,
                field_type=None,  # Type inference deferred
                is_required=col_idx <= _REQUIRED_COLUMN_COUNT,
            )
        )

    wb.close()

    # Derive version label from filename
    version_label = filename.rsplit(".", 1)[0] if "." in filename else filename

    logger.info(
        f"Parsed BAR template '{filename}': "
        f"{len(columns)} columns from row {header_row_idx}"
    )

    return BARTemplate(
        filename=filename,
        version_label=version_label,
        column_count=len(columns),
        columns=columns,
        raw_header_row=raw_header,
    )


async def get_active_template() -> Optional[BARTemplate]:
    """Get the currently active BAR template."""
    return await BARTemplate.get_active()
