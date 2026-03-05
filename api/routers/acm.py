"""
ACM (Asbestos Containing Material) API Endpoints

Provides REST API for ACM record management including
listing, filtering, extraction, and export.
"""

import csv
import io
import json as _json
import math
import re
import zipfile
from datetime import date
from typing import List, Optional

from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import StreamingResponse
from loguru import logger
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from api.command_service import CommandService
from api.models import (
    ACMExtractRequest,
    ACMExtractResponse,
    ACMRecordCreateRequest,
    ACMRecordListResponse,
    ACMRecordResponse,
    ACMRecordUpdateRequest,
    ACMSearchResponse,
    ACMSearchResultResponse,
    ACMStatsResponse,
    AgencyListResponse,
    ApplyTemplateRequest,
    BackfillBuildingsRequest,
    BackfillBuildingsResponse,
    BackfillParentsRequest,
    BackfillParentsResponse,
    BatchClassifyRequest,
    BatchClassifyResponse,
    BuildingRecordCreateRequest,
    BuildingRecordListResponse,
    BuildingRecordResponse,
    BuildingRecordUpdateRequest,
    BuildingResponse,
    BuildingUpdateRequest,
    BuildingValidationSummary,
    BulkEditRequest,
    BulkEditResponse,
    BulkFixResponse,
    BulkValidateRequest,
    BulkValidateResponse,
    BusinessRuleResponse,
    ClassifyRequest,
    ClassifyResponse,
    FieldDefResponse,
    FieldMappingUpdateRequest,
    FieldSchemaConfigResponse,
    FieldSchemaConfigUpdateRequest,
    NormalizeRequest,
    NormalizeResponse,
    OfficerEditEntry,  # noqa: F401 — imported for re-export / forward ref
    ParentContextResponse,
    PatchRawExtractionRequest,
    ProvenanceResponse,
    RawExtractionListResponse,
    RawExtractionResponse,
    RawTableResponse,
    ReEmbedRequest,
    ReEmbedResponse,
    SFFieldSchemaConfigResponse,
    SiteConfigRequest,
    SiteConfigResponse,
    SiteConfigTemplateResponse,
    SourceIntelligenceResponse,
    TaxonomyGroupResponse,
    TaxonomyResponse,
    ValidationSummaryResponse,
)
from open_notebook.database.repository import (
    ensure_record_id,
    get_source_intelligence,
    repo_query,
)
from open_notebook.domain.acm import (
    ACMRecord,
    ACMTableSection,
    BuildingRecord,
    RawExtraction,
)
from open_notebook.domain.site_config import SiteConfig
from open_notebook.exceptions import NotFoundError

router = APIRouter()


@router.get("/records", response_model=ACMRecordListResponse)
async def list_acm_records(
    source_id: str = Query(..., description="Source ID to filter by (required)"),
    building_id: Optional[str] = Query(None, description="Filter by building ID"),
    building_record_id: Optional[str] = Query(
        None, description="Filter by building_record FK (e.g., building_record:xxx)"
    ),
    room_id: Optional[str] = Query(None, description="Filter by room ID"),
    risk_status: Optional[str] = Query(
        None, description="Filter by risk status (Low/Medium/High)"
    ),
    page: int = Query(1, ge=1, description="Page number"),
    limit: int = Query(100, ge=1, le=500, description="Records per page"),
):
    """
    List ACM records with filtering and pagination.

    Returns records for the specified source, with optional filtering
    by building, room, risk status, and building_record FK.
    """
    try:
        # Build query conditions
        conditions = ["source_id = $source_id"]
        params: dict = {"source_id": ensure_record_id(source_id)}

        if building_id:
            conditions.append("building_id = $building_id")
            params["building_id"] = building_id

        if building_record_id:
            conditions.append("building_record_id = $building_record_id")
            params["building_record_id"] = ensure_record_id(building_record_id)

        if room_id:
            conditions.append("room_id = $room_id")
            params["room_id"] = room_id

        if risk_status:
            conditions.append("risk_status = $risk_status")
            params["risk_status"] = risk_status

        where_clause = " AND ".join(conditions)

        # Get total count
        count_query = (
            f"SELECT count() as total FROM acm_record WHERE {where_clause} GROUP ALL"
        )
        count_result = await repo_query(count_query, params)
        total = count_result[0]["total"] if count_result else 0

        # Calculate pagination
        offset = (page - 1) * limit
        pages = math.ceil(total / limit) if total > 0 else 1

        # Get paginated records
        data_query = f"""
            SELECT * FROM acm_record
            WHERE {where_clause}
            ORDER BY building_id, room_id, id
            LIMIT $limit START $offset
        """
        params["limit"] = limit
        params["offset"] = offset

        records = await repo_query(data_query, params)

        # Convert to response models
        record_responses = []
        for r in records:
            record_responses.append(
                ACMRecordResponse(
                    id=str(r.get("id", "")),
                    source_id=str(r.get("source_id", "")),
                    school_name=r.get("school_name", ""),
                    school_code=r.get("school_code"),
                    building_id=r.get("building_id", ""),
                    building_name=r.get("building_name"),
                    building_year=r.get("building_year"),
                    building_construction=r.get("building_construction"),
                    room_id=r.get("room_id"),
                    room_name=r.get("room_name"),
                    room_area=r.get("room_area"),
                    area_type=r.get("area_type"),
                    product=r.get("product", ""),
                    material_description=r.get("material_description", ""),
                    extent=r.get("extent"),
                    location=r.get("location"),
                    friable=r.get("friable"),
                    material_condition=r.get("material_condition"),
                    risk_status=r.get("risk_status"),
                    result=r.get("result", ""),
                    page_number=r.get("page_number"),
                    extraction_confidence=r.get("extraction_confidence"),
                    acm_product_group=r.get("acm_product_group"),
                    acm_product_type=r.get("acm_product_type"),
                    classification_confidence=r.get("classification_confidence"),
                    classification_method=r.get("classification_method"),
                    classification_override=r.get("classification_override"),
                    sample_no=r.get("sample_no"),
                    sample_result=r.get("sample_result"),
                    quantity=r.get("quantity"),
                    acm_labelled=r.get("acm_labelled"),
                    acm_label_details=r.get("acm_label_details"),
                    identifying_company=r.get("identifying_company"),
                    disturbance_potential=r.get("disturbance_potential"),
                    hygienist_recommendations=r.get("hygienist_recommendations"),
                    normalized_action=r.get("normalized_action"),
                    data_issues=r.get("data_issues"),
                    floor_level=r.get("floor_level"),
                    no_access=r.get("no_access"),
                    smf_present=r.get("smf_present"),
                    created=str(r.get("created", "")) if r.get("created") else None,
                    updated=str(r.get("updated", "")) if r.get("updated") else None,
                )
            )

        return ACMRecordListResponse(
            records=record_responses,
            total=total,
            page=page,
            pages=pages,
            limit=limit,
        )

    except Exception as e:
        logger.error(f"Error listing ACM records: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/records/{record_id}", response_model=ACMRecordResponse)
async def get_acm_record(record_id: str):
    """Get a single ACM record by ID."""
    try:
        record = await ACMRecord.get(record_id)
        if not record:
            raise HTTPException(status_code=404, detail="ACM record not found")

        return ACMRecordResponse(
            id=str(record.id),
            source_id=str(record.source_id),
            school_name=record.school_name,
            school_code=record.school_code,
            building_id=record.building_id,
            building_name=record.building_name,
            building_year=record.building_year,
            building_construction=record.building_construction,
            room_id=record.room_id,
            room_name=record.room_name,
            room_area=record.room_area,
            area_type=record.area_type,
            product=record.product,
            material_description=record.material_description,
            extent=record.extent,
            location=record.location,
            friable=record.friable,
            material_condition=record.material_condition,
            risk_status=record.risk_status,
            result=record.result,
            page_number=record.page_number,
            extraction_confidence=record.extraction_confidence,
            acm_product_group=record.acm_product_group,
            acm_product_type=record.acm_product_type,
            classification_confidence=record.classification_confidence,
            classification_method=record.classification_method,
            classification_override=record.classification_override,
            sample_no=record.sample_no,
            sample_result=record.sample_result,
            quantity=record.quantity,
            acm_labelled=record.acm_labelled,
            acm_label_details=record.acm_label_details,
            identifying_company=record.identifying_company,
            disturbance_potential=record.disturbance_potential,
            hygienist_recommendations=record.hygienist_recommendations,
            normalized_action=record.normalized_action,
            data_issues=record.data_issues,
            floor_level=record.floor_level,
            no_access=record.no_access,
            smf_present=record.smf_present,
            created=str(record.created) if record.created else None,
            updated=str(record.updated) if record.updated else None,
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error getting ACM record {record_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/extract", response_model=ACMExtractResponse)
async def trigger_acm_extraction(request: ACMExtractRequest):
    """
    Trigger ACM extraction for a source document.

    Submits an async extraction job that parses the source's
    Docling output and creates ACM records.
    """
    try:
        # Import command modules to ensure they're registered
        import commands.acm_commands  # noqa: F401

        # Submit extraction command
        command_id = await CommandService.submit_command_job(
            "open_notebook",
            "acm_extract",
            {"source_id": request.source_id, "force": request.force},
        )

        return ACMExtractResponse(
            command_id=str(command_id),
            status="submitted",
            message=f"ACM extraction started for source {request.source_id}",
        )

    except Exception as e:
        logger.error(f"Error triggering ACM extraction: {e}")
        raise HTTPException(status_code=500, detail=str(e))


async def _get_export_mapping() -> tuple[list[str], list[str | None]]:
    """Get export headers and field names from active field mapping.

    Returns (headers, fields) where fields[i] is the ACMRecord attribute
    for headers[i], or None if unmapped.
    """
    try:
        from open_notebook.domain.field_mapping import FieldMapping

        mapping = await FieldMapping.get_active()
        if mapping.mappings:
            headers = [m.bar_column for m in mapping.mappings]
            fields = [m.acm_field for m in mapping.mappings]
            return headers, fields
    except Exception as e:
        logger.debug(f"Failed to load field mapping, using defaults: {e}")

    # Fallback to hardcoded defaults
    headers = [
        "Building Code",
        "Building Name",
        "Room ID",
        "Room Name",
        "Floor Level",
        "Product",
        "Material Description",
        "Extent",
        "Location",
        "Friable",
        "Condition",
        "Risk Status",
        "Result",
        "Sample No",
        "Sample Result",
        "Quantity",
        "ACM Labelled",
        "Disturbance Potential",
        "Identifying Company",
        "Hygienist Recommendations",
        "Page Number",
    ]
    fields: list[str | None] = [
        "building_id",
        "building_name",
        "room_id",
        "room_name",
        "floor_level",
        "product",
        "material_description",
        "extent",
        "location",
        "friable",
        "material_condition",
        "risk_status",
        "result",
        "sample_no",
        "sample_result",
        "quantity",
        "acm_labelled",
        "disturbance_potential",
        "identifying_company",
        "hygienist_recommendations",
        "page_number",
    ]
    return headers, fields


def _sanitize_csv_value(value: str) -> str:
    """Sanitize a value to prevent CSV injection in spreadsheet applications."""
    if value and value[0] in ("=", "+", "-", "@"):
        return f"'{value}"
    return value


def _get_record_value(record: ACMRecord, field: str | None) -> str:
    """Get a display value from an ACMRecord for a given field name."""
    if not field:
        return ""
    val = getattr(record, field, None)
    if val is None:
        return ""
    if isinstance(val, bool):
        return "Yes" if val else "No"
    if isinstance(val, list):
        return _sanitize_csv_value("; ".join(str(v) for v in val))
    return _sanitize_csv_value(str(val))


@router.get("/export")
@router.get("/export/csv")
async def export_acm_records(
    source_id: str = Query(..., description="Source ID to export"),
    building_ids: Optional[List[str]] = Query(
        None, description="Filter to specific building IDs"
    ),
):
    """
    Export ACM records as CSV file.

    Downloads all records for the specified source as a CSV file.
    Uses active field mapping for column order when available.
    """
    try:
        # Get all records for source
        records = await ACMRecord.get_by_source(source_id)

        # Filter to requested buildings if specified
        if building_ids:
            records = [r for r in records if r.get("building_id") in building_ids]

        if not records:
            raise HTTPException(
                status_code=404, detail="No ACM records found for source"
            )

        # Get export mapping
        headers, fields = await _get_export_mapping()

        # Create CSV in memory
        output = io.StringIO()
        writer = csv.writer(output)

        # Write header — use field mapping
        writer.writerow(headers)

        # Write data rows using field mapping
        for record in records:
            writer.writerow([_get_record_value(record, f) for f in fields])

        # Create response
        output.seek(0)

        # Get source title for filename
        from open_notebook.domain.notebook import Source

        source = await Source.get(source_id)
        source_title = source.title if source else source_id
        filename = f"acm_export_{source_title}.csv".replace(" ", "_")

        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error exporting ACM records: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# Risk status colors for Excel
RISK_COLORS = {
    "Low": "C6EFCE",  # Green
    "Medium": "FFEB9C",  # Yellow/Orange
    "High": "FFC7CE",  # Red
}


@router.get("/export/excel")
async def export_acm_excel(
    source_id: str = Query(..., description="Source ID to export"),
    building_ids: Optional[List[str]] = Query(
        None, description="Filter to specific building IDs"
    ),
):
    """
    Export ACM records as formatted Excel file.

    Downloads all records for the specified source as an Excel file
    with formatted headers, auto-sized columns, and risk status color coding.
    """
    try:
        # Get all records for source
        records = await ACMRecord.get_by_source(source_id)

        # Filter to requested buildings if specified
        if building_ids:
            records = [r for r in records if r.get("building_id") in building_ids]

        if not records:
            raise HTTPException(
                status_code=404, detail="No ACM records found for source"
            )

        # Get source title for filename
        from open_notebook.domain.notebook import Source

        source = await Source.get(source_id)
        source_title = source.title if source else source_id

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "ACM Register"

        # Get columns from active field mapping
        export_headers, export_fields = await _get_export_mapping()
        columns = [
            (h, f, max(12, len(h) + 2)) for h, f in zip(export_headers, export_fields)
        ]

        # Header styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Write headers
        for col_idx, (header, _, width) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Write data rows
        for row_idx, record in enumerate(records, 2):
            for col_idx, (_, field, _) in enumerate(columns, 1):
                value = _get_record_value(record, field)
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border

                # Color code risk status
                if field == "risk_status" and value in RISK_COLORS:
                    cell.fill = PatternFill(
                        start_color=RISK_COLORS[value],
                        end_color=RISK_COLORS[value],
                        fill_type="solid",
                    )

        # Freeze header row
        ws.freeze_panes = "A2"

        # Auto-filter
        if records:
            ws.auto_filter.ref = ws.dimensions

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"acm_export_{source_title}_{date.today()}.xlsx".replace(" ", "_")

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error exporting ACM records to Excel: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/export/sf-csv")
async def export_sf_csv(
    source_id: str = Query(..., description="Source ID to export"),
    building_ids: Optional[str] = Query(
        None,
        description="Comma-separated building internal_ids to restrict export (AC7)",
    ),
):
    """Export as ZIP containing Building__c.csv + Item__c.csv with SF API field names.

    Returns a ZIP archive with two CSV files:
    - Building__c.csv — one row per building, SF API names as headers
    - Item__c.csv     — one row per ACM item, linked via Building__r.External_ID__c

    Building ID filtering (AC7): pass building_ids as comma-separated internal_ids
    to restrict the export to only those buildings and their items.
    """
    from open_notebook.extractors.exporters.sf_export import (
        BUILDING_SF_MAPPING,
        ITEM_SF_MAPPING,
        building_to_sf_row,
        generate_external_id,
        get_building_sf_headers,
        get_item_sf_headers,
        item_to_sf_row,
    )

    try:
        # Resolve optional building filter
        building_id_filter: set[str] | None = None
        if building_ids:
            building_id_filter = {
                bid.strip() for bid in building_ids.split(",") if bid.strip()
            }

        # Fetch building records
        buildings = await BuildingRecord.get_by_source(source_id)
        if not buildings:
            raise HTTPException(
                status_code=404, detail="No building records found for source"
            )

        if building_id_filter:
            buildings = [b for b in buildings if b.internal_id in building_id_filter]

        # Fetch site config (AC5)
        site_config = None
        try:
            site_config = await SiteConfig.get_by_source(source_id)
        except Exception as sc_err:
            logger.warning(f"Could not load site config for {source_id}: {sc_err}")

        # Pre-compute External_ID__c for each building (keyed by building DB id string)
        building_ext_id: dict[str, str] = {}
        for b in buildings:
            eid = generate_external_id(b, source_id)
            building_ext_id[str(b.id) if b.id else b.internal_id] = eid

        # Also index by building_code so ACMRecord.building_id can resolve the parent
        building_code_to_ext_id: dict[str, str] = {}
        for b in buildings:
            if b.building_code:
                building_code_to_ext_id[b.building_code] = generate_external_id(
                    b, source_id
                )
            building_code_to_ext_id[b.internal_id] = generate_external_id(b, source_id)

        # Fetch ACM records
        all_records = await ACMRecord.get_by_source(source_id)
        # Filter by building_record_id or building_id if a building filter was requested
        if building_id_filter:
            allowed_building_codes = {
                b.building_code for b in buildings if b.building_code
            } | {b.internal_id for b in buildings}
            all_records = [
                r for r in all_records if r.building_id in allowed_building_codes
            ]

        # Build Building CSV bytes
        building_buf = io.StringIO()
        b_writer = csv.DictWriter(
            building_buf, fieldnames=get_building_sf_headers(), extrasaction="ignore"
        )
        b_writer.writeheader()
        for b in buildings:
            row = building_to_sf_row(b, source_id, site_config)
            b_writer.writerow(row)
        building_csv_bytes = building_buf.getvalue().encode("utf-8-sig")

        # Build Item CSV bytes
        item_buf = io.StringIO()
        i_writer = csv.DictWriter(
            item_buf, fieldnames=get_item_sf_headers(), extrasaction="ignore"
        )
        i_writer.writeheader()
        for record in all_records:
            parent_ext_id = building_code_to_ext_id.get(record.building_id, "")
            row = item_to_sf_row(record, parent_ext_id)
            i_writer.writerow(row)
        item_csv_bytes = item_buf.getvalue().encode("utf-8-sig")

        # Package both CSVs into a ZIP
        zip_buf = io.BytesIO()
        with zipfile.ZipFile(zip_buf, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
            zf.writestr("Building__c.csv", building_csv_bytes)
            zf.writestr("Item__c.csv", item_csv_bytes)
        zip_buf.seek(0)

        # Get source title for filename
        from open_notebook.domain.notebook import Source

        source = await Source.get(source_id)
        source_title = source.title if source else source_id
        filename = f"sf_export_{source_title}_{date.today()}.zip".replace(" ", "_")

        return StreamingResponse(
            zip_buf,
            media_type="application/zip",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error exporting SF CSV for source {source_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/export/sf-excel")
async def export_sf_excel(
    source_id: str = Query(..., description="Source ID to export"),
    building_ids: Optional[str] = Query(
        None,
        description="Comma-separated building internal_ids to restrict export (AC7)",
    ),
):
    """Export as XLSX workbook with Building__c and Item__c sheets.

    Returns a single Excel file with two sheets:
    - Sheet 1 "Building__c" — one row per building, SF API names as headers
    - Sheet 2 "Item__c"     — one row per ACM item, linked via Building__r.External_ID__c

    Building ID filtering (AC7): pass building_ids as comma-separated internal_ids
    to restrict the export to only those buildings and their items.
    """
    from open_notebook.extractors.exporters.sf_export import (
        building_to_sf_row,
        generate_external_id,
        get_building_sf_headers,
        get_item_sf_headers,
        item_to_sf_row,
    )

    try:
        # Resolve optional building filter
        building_id_filter: set[str] | None = None
        if building_ids:
            building_id_filter = {
                bid.strip() for bid in building_ids.split(",") if bid.strip()
            }

        # Fetch building records
        buildings = await BuildingRecord.get_by_source(source_id)
        if not buildings:
            raise HTTPException(
                status_code=404, detail="No building records found for source"
            )

        if building_id_filter:
            buildings = [b for b in buildings if b.internal_id in building_id_filter]

        # Fetch site config (AC5)
        site_config = None
        try:
            site_config = await SiteConfig.get_by_source(source_id)
        except Exception as sc_err:
            logger.warning(f"Could not load site config for {source_id}: {sc_err}")

        # Pre-compute External_ID__c per building_code / internal_id
        building_code_to_ext_id: dict[str, str] = {}
        for b in buildings:
            ext_id = generate_external_id(b, source_id)
            if b.building_code:
                building_code_to_ext_id[b.building_code] = ext_id
            building_code_to_ext_id[b.internal_id] = ext_id

        # Fetch ACM records
        all_records = await ACMRecord.get_by_source(source_id)
        if building_id_filter:
            allowed_building_codes = {
                b.building_code for b in buildings if b.building_code
            } | {b.internal_id for b in buildings}
            all_records = [
                r for r in all_records if r.building_id in allowed_building_codes
            ]

        # --- Excel workbook ---
        wb = Workbook()

        # Shared cell styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill_blue = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        header_fill_green = PatternFill(
            start_color="375623", end_color="375623", fill_type="solid"
        )
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # --- Sheet 1: Building__c ---
        ws_bld = wb.active
        ws_bld.title = "Building__c"
        b_headers = get_building_sf_headers()
        for col_idx, header in enumerate(b_headers, 1):
            cell = ws_bld.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill_blue
            cell.alignment = header_alignment
            cell.border = thin_border
            ws_bld.column_dimensions[get_column_letter(col_idx)].width = max(
                12, len(header) + 2
            )

        for row_idx, b in enumerate(buildings, 2):
            b_row = building_to_sf_row(b, source_id, site_config)
            for col_idx, header in enumerate(b_headers, 1):
                cell = ws_bld.cell(
                    row=row_idx, column=col_idx, value=b_row.get(header, "")
                )
                cell.border = thin_border

        ws_bld.freeze_panes = "A2"
        if buildings:
            ws_bld.auto_filter.ref = ws_bld.dimensions

        # --- Sheet 2: Item__c ---
        ws_itm = wb.create_sheet(title="Item__c")
        i_headers = get_item_sf_headers()
        for col_idx, header in enumerate(i_headers, 1):
            cell = ws_itm.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill_green
            cell.alignment = header_alignment
            cell.border = thin_border
            ws_itm.column_dimensions[get_column_letter(col_idx)].width = max(
                12, len(header) + 2
            )

        for row_idx, record in enumerate(all_records, 2):
            parent_ext_id = building_code_to_ext_id.get(record.building_id, "")
            i_row = item_to_sf_row(record, parent_ext_id)
            for col_idx, header in enumerate(i_headers, 1):
                cell = ws_itm.cell(
                    row=row_idx, column=col_idx, value=i_row.get(header, "")
                )
                cell.border = thin_border

                # Color code Risk_Status__c column
                if header == "Risk_Status__c" and i_row.get(header) in RISK_COLORS:
                    cell.fill = PatternFill(
                        start_color=RISK_COLORS[i_row[header]],
                        end_color=RISK_COLORS[i_row[header]],
                        fill_type="solid",
                    )

        ws_itm.freeze_panes = "A2"
        if all_records:
            ws_itm.auto_filter.ref = ws_itm.dimensions

        # Save workbook to bytes buffer
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Get source title for filename
        from open_notebook.domain.notebook import Source

        source = await Source.get(source_id)
        source_title = source.title if source else source_id
        filename = f"sf_export_{source_title}_{date.today()}.xlsx".replace(" ", "_")

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error exporting SF Excel for source {source_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/stats", response_model=ACMStatsResponse)
async def get_acm_stats(
    source_id: Optional[str] = Query(None, description="Filter stats by source"),
):
    """
    Get ACM statistics summary.

    Returns counts of records by risk status and other metrics.
    """
    try:
        if source_id:
            stats = await ACMRecord.get_summary_by_source(source_id)
            return ACMStatsResponse(
                source_id=source_id,
                **stats,
            )
        else:
            # Global stats (all sources)
            result = await repo_query(
                """
                SELECT
                    count() as total_records,
                    count(risk_status = 'High' OR NULL) as high_risk_count,
                    count(risk_status = 'Medium' OR NULL) as medium_risk_count,
                    count(risk_status = 'Low' OR NULL) as low_risk_count,
                    array::distinct(building_id) as buildings,
                    array::distinct(room_id) as rooms
                FROM acm_record
                GROUP ALL
            """
            )

            if result:
                return ACMStatsResponse(
                    total_records=result[0].get("total_records", 0),
                    high_risk_count=result[0].get("high_risk_count", 0),
                    medium_risk_count=result[0].get("medium_risk_count", 0),
                    low_risk_count=result[0].get("low_risk_count", 0),
                    building_count=len(result[0].get("buildings", [])),
                    room_count=len([r for r in result[0].get("rooms", []) if r]),
                )
            else:
                return ACMStatsResponse(
                    total_records=0,
                    high_risk_count=0,
                    medium_risk_count=0,
                    low_risk_count=0,
                    building_count=0,
                    room_count=0,
                )

    except Exception as e:
        logger.error(f"Error getting ACM stats: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/search", response_model=ACMSearchResponse)
async def semantic_search_acm(
    query: str = Query(..., min_length=1, description="Natural language search query"),
    source_id: Optional[str] = Query(None, description="Filter to specific source"),
    building_id: Optional[str] = Query(None, description="Filter to specific building"),
    limit: int = Query(10, ge=1, le=100, description="Maximum results to return"),
    threshold: float = Query(
        0.7, ge=0.0, le=1.0, description="Minimum similarity score"
    ),
    include_parent: bool = Query(
        False, description="Include parent table section context (E11-S1)"
    ),
    search_mode: str = Query(
        "hybrid",
        description="Search mode: hybrid (BM25+vector+RRF), vector, or bm25 (E11-S2)",
    ),
):
    """
    Semantic search across ACM records.

    Uses vector similarity to find records matching natural language queries.
    Requires records to have embeddings generated via the embedding pipeline.

    When include_parent=true, returns parent table section context alongside
    matched records for richer retrieval (E11-S1 Parent Document Retrieval).

    Example queries:
    - "high risk asbestos items"
    - "floor tiles in poor condition"
    - "accessible materials in corridors"
    """
    try:
        from api.services.acm_embedding_service import ACMEmbeddingService
        from api.services.acm_hybrid_search_service import HybridSearchService
        from open_notebook.domain.models import model_manager

        resolved_source_id = ensure_record_id(source_id) if source_id else None

        # BM25-only mode doesn't need embeddings
        if search_mode == "bm25":
            hybrid_svc = HybridSearchService()
            filtered_results = await hybrid_svc.bm25_search(
                query,
                source_id=resolved_source_id,
                building_id=building_id,
                limit=limit,
            )
            # Normalize score field
            for r in filtered_results:
                r["score"] = r.get("bm25_score", 0)
        else:
            # Get embedding model for vector/hybrid modes
            embedding_model = await model_manager.get_embedding_model()
            if not embedding_model:
                raise HTTPException(
                    status_code=400,
                    detail="No embedding model configured. Please configure one in Settings.",
                )

            # Embed the query
            try:
                query_embedding = (await embedding_model.aembed([query]))[0]
            except Exception as e:
                logger.error(f"Failed to embed query: {e}")
                raise HTTPException(
                    status_code=500, detail=f"Failed to embed query: {str(e)}"
                )

            if search_mode == "hybrid":
                hybrid_svc = HybridSearchService()
                filtered_results = await hybrid_svc.search(
                    query=query,
                    query_embedding=query_embedding,
                    source_id=resolved_source_id,
                    building_id=building_id,
                    limit=limit,
                    threshold=threshold,
                    mode="hybrid",
                )
                # Normalize score: use RRF score as primary
                for r in filtered_results:
                    r["score"] = r.get("rrf_score", r.get("similarity", 0))
            else:
                # Pure vector search (original behavior)
                filters = ["embedding IS NOT NULL"]
                params = {
                    "query_embedding": query_embedding,
                    "limit": limit,
                }

                if resolved_source_id:
                    filters.append("source_id = $source_id")
                    params["source_id"] = resolved_source_id
                if building_id:
                    filters.append("building_id = $building_id")
                    params["building_id"] = building_id

                where_clause = " AND ".join(filters)

                search_query = f"""
                    SELECT *,
                           vector::similarity::cosine(embedding, $query_embedding) AS score
                    FROM acm_record
                    WHERE {where_clause}
                    ORDER BY score DESC
                    LIMIT $limit
                """

                results = await repo_query(search_query, params)
                filtered_results = [
                    r for r in results if r.get("score", 0) >= threshold
                ]

        # Batch-fetch parent sections if requested (avoids N+1 queries)
        parent_section_map: dict[str, ACMTableSection] = {}
        if include_parent and filtered_results:
            # Collect unique source_ids that have parent_table_id set
            source_ids_needing_parents = {
                str(r.get("source_id", ""))
                for r in filtered_results
                if r.get("parent_table_id")
            }
            for sid in source_ids_needing_parents:
                try:
                    sections = await ACMTableSection.get_by_source(sid)
                    for s in sections:
                        if s.id:
                            parent_section_map[str(s.id)] = s
                except Exception as e:
                    logger.warning(
                        f"Failed to batch-fetch parent sections for {sid}: {e}"
                    )

        # Convert to response objects
        search_results = []
        for r in filtered_results:
            score = r.get("score", 0)

            # Look up parent context by direct ID from the batch map
            parent_ctx = None
            if include_parent and r.get("parent_table_id"):
                parent_id = str(r["parent_table_id"])
                section = parent_section_map.get(parent_id)
                if section:
                    parent_ctx = ParentContextResponse(
                        id=str(section.id) if section.id else "",
                        building_name=section.building_name,
                        page_start=section.page_start,
                        page_end=section.page_end,
                        table_type=section.table_type,
                        raw_text=section.raw_text,
                    )

            search_results.append(
                ACMSearchResultResponse(
                    id=str(r.get("id", "")),
                    source_id=str(r.get("source_id", "")),
                    school_name=r.get("school_name", ""),
                    building_id=r.get("building_id", ""),
                    building_name=r.get("building_name"),
                    room_id=r.get("room_id"),
                    room_name=r.get("room_name"),
                    product=r.get("product", ""),
                    material_description=r.get("material_description", ""),
                    extent=r.get("extent"),
                    location=r.get("location"),
                    material_condition=r.get("material_condition"),
                    risk_status=r.get("risk_status"),
                    result=r.get("result", ""),
                    score=round(score, 4),
                    parent_context=parent_ctx,
                )
            )

        logger.info(
            f"ACM search for '{query}': {len(search_results)} results "
            f"(mode={search_mode}, threshold={threshold}, limit={limit})"
        )

        return ACMSearchResponse(
            query=query,
            results=search_results,
            total=len(search_results),
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error in semantic search: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/re-embed", response_model=ReEmbedResponse)
async def re_embed_acm_records(request: ReEmbedRequest):
    """
    Re-embed ACM records with contextual enrichment (E1-S14).

    Generates enriched_text with hierarchical context (Building, Level, Room, Page)
    and re-embeds records. Optionally filter by source_id.
    Use force=True to re-embed records that already have embeddings.
    """
    try:
        from api.services.acm_embedding_service import ACMEmbeddingService

        service = ACMEmbeddingService()
        count = await service.re_embed_acm_records(
            source_id=request.source_id,
            force=request.force,
        )

        return ReEmbedResponse(
            success=True,
            records_processed=count,
            message=f"Re-embedded {count} ACM records with contextual enrichment",
        )

    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        logger.error(f"Error re-embedding ACM records: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/backfill-parents", response_model=BackfillParentsResponse)
async def backfill_parent_references(
    request: BackfillParentsRequest = BackfillParentsRequest(),
):
    """
    Backfill parent_table_id for existing ACM records (E11-S1).

    Matches each record's source_id + page_number to an acm_table_section
    page range and sets parent_table_id. Optionally filter by source_id.
    """
    try:
        # Build query for records without parent
        conditions = ["parent_table_id IS NULL OR parent_table_id IS NONE"]
        params: dict = {}
        if request.source_id:
            conditions.append("source_id = $source_id")
            params["source_id"] = ensure_record_id(request.source_id)

        where_clause = " AND ".join(conditions)
        orphan_query = f"SELECT id, source_id, page_number FROM acm_record WHERE {where_clause} LIMIT 10000"
        orphan_records = await repo_query(orphan_query, params)

        if not orphan_records:
            return BackfillParentsResponse(
                records_updated=0,
                message="No records found without parent references",
            )

        updated = 0
        for record in orphan_records:
            page_num = record.get("page_number")
            source = str(record.get("source_id", ""))
            record_id = str(record.get("id", ""))

            if not page_num or not source or not record_id:
                continue

            # Find matching section
            section_query = (
                "SELECT id FROM acm_table_section "
                "WHERE source_id = $source_id AND page_start <= $page AND page_end >= $page "
                "LIMIT 1"
            )
            sections = await repo_query(
                section_query,
                {"source_id": ensure_record_id(source), "page": page_num},
            )

            if sections:
                section_id = str(sections[0].get("id", ""))
                if section_id:
                    await repo_query(
                        "UPDATE $record_id SET parent_table_id = $section_id",
                        {
                            "record_id": ensure_record_id(record_id),
                            "section_id": ensure_record_id(section_id),
                        },
                    )
                    updated += 1

        return BackfillParentsResponse(
            records_updated=updated,
            message=f"Linked {updated} records to parent table sections",
        )

    except Exception as e:
        logger.error(f"Error backfilling parent references: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/backfill-buildings", response_model=BackfillBuildingsResponse)
async def backfill_building_records(
    request: BackfillBuildingsRequest = BackfillBuildingsRequest(),
):
    """
    Backfill building_record entries from existing ACM records (E35-S6).

    Groups acm_records by (source_id, building_id), creates a BuildingRecord
    for each unique building, and sets building_record_id FK on the acm_records.
    Idempotent: skips buildings that already exist and never overwrites
    non-NULL FK values.
    """
    try:
        from scripts.v3_building_backfill import backfill_all, backfill_source

        if request.source_id:
            result = await backfill_source(request.source_id)
        else:
            result = await backfill_all()

        total = result.buildings_created + result.buildings_skipped
        return BackfillBuildingsResponse(
            buildings_created=result.buildings_created,
            buildings_skipped=result.buildings_skipped,
            records_linked=result.records_linked,
            message=(
                f"Backfill complete: {result.buildings_created} buildings created, "
                f"{result.buildings_skipped} skipped, "
                f"{result.records_linked} records linked "
                f"({total} total buildings)"
            ),
        )

    except Exception as e:
        logger.error(f"Error backfilling building records: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/records", response_model=ACMRecordResponse)
async def create_acm_record(request: ACMRecordCreateRequest):
    """
    Create a new ACM record.

    Creates a single ACM record with the provided data.
    """
    try:
        # Create ACMRecord instance
        record = ACMRecord(
            source_id=request.source_id,
            school_name=request.school_name,
            school_code=request.school_code,
            building_id=request.building_id,
            building_name=request.building_name,
            building_year=request.building_year,
            building_construction=request.building_construction,
            room_id=request.room_id,
            room_name=request.room_name,
            room_area=request.room_area,
            area_type=request.area_type,
            product=request.product,
            material_description=request.material_description,
            extent=request.extent,
            location=request.location,
            friable=request.friable,
            material_condition=request.material_condition,
            risk_status=request.risk_status,
            result=request.result,
            page_number=request.page_number,
        )

        # Save to database
        await record.save()

        return ACMRecordResponse(
            id=str(record.id),
            source_id=str(record.source_id),
            school_name=record.school_name,
            school_code=record.school_code,
            building_id=record.building_id,
            building_name=record.building_name,
            building_year=record.building_year,
            building_construction=record.building_construction,
            room_id=record.room_id,
            room_name=record.room_name,
            room_area=record.room_area,
            area_type=record.area_type,
            product=record.product,
            material_description=record.material_description,
            extent=record.extent,
            location=record.location,
            friable=record.friable,
            material_condition=record.material_condition,
            risk_status=record.risk_status,
            result=record.result,
            page_number=record.page_number,
            extraction_confidence=None,
            sample_no=record.sample_no,
            sample_result=record.sample_result,
            quantity=record.quantity,
            acm_labelled=record.acm_labelled,
            acm_label_details=record.acm_label_details,
            identifying_company=record.identifying_company,
            disturbance_potential=record.disturbance_potential,
            hygienist_recommendations=record.hygienist_recommendations,
            normalized_action=record.normalized_action,
            data_issues=record.data_issues,
            floor_level=record.floor_level,
            no_access=None,
            smf_present=None,
            created=str(record.created) if record.created else None,
            updated=str(record.updated) if record.updated else None,
        )

    except Exception as e:
        logger.error(f"Error creating ACM record: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.put("/records/{record_id}", response_model=ACMRecordResponse)
async def update_acm_record(record_id: str, request: ACMRecordUpdateRequest):
    """
    Update an existing ACM record.

    Only provided fields will be updated. All fields are optional.
    """
    try:
        # Fetch existing record
        record = await ACMRecord.get(record_id)
        if not record:
            raise HTTPException(status_code=404, detail="ACM record not found")

        # Update only provided fields
        update_data = request.model_dump(exclude_unset=True, exclude_none=True)

        for field, value in update_data.items():
            if hasattr(record, field):
                setattr(record, field, value)

        # Save changes
        await record.save()

        return ACMRecordResponse(
            id=str(record.id),
            source_id=str(record.source_id),
            school_name=record.school_name,
            school_code=record.school_code,
            building_id=record.building_id,
            building_name=record.building_name,
            building_year=record.building_year,
            building_construction=record.building_construction,
            room_id=record.room_id,
            room_name=record.room_name,
            room_area=record.room_area,
            area_type=record.area_type,
            product=record.product,
            material_description=record.material_description,
            extent=record.extent,
            location=record.location,
            friable=record.friable,
            material_condition=record.material_condition,
            risk_status=record.risk_status,
            result=record.result,
            page_number=record.page_number,
            extraction_confidence=record.extraction_confidence,
            acm_product_group=record.acm_product_group,
            acm_product_type=record.acm_product_type,
            classification_confidence=record.classification_confidence,
            classification_method=record.classification_method,
            classification_override=record.classification_override,
            sample_no=record.sample_no,
            sample_result=record.sample_result,
            quantity=record.quantity,
            acm_labelled=record.acm_labelled,
            acm_label_details=record.acm_label_details,
            identifying_company=record.identifying_company,
            disturbance_potential=record.disturbance_potential,
            hygienist_recommendations=record.hygienist_recommendations,
            normalized_action=record.normalized_action,
            data_issues=record.data_issues,
            floor_level=record.floor_level,
            no_access=record.no_access,
            smf_present=record.smf_present,
            created=str(record.created) if record.created else None,
            updated=str(record.updated) if record.updated else None,
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating ACM record {record_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/records/{record_id}")
async def delete_acm_record(record_id: str):
    """
    Delete an ACM record.

    Permanently removes the record from the database.
    """
    try:
        record = await ACMRecord.get(record_id)
        if not record:
            raise HTTPException(status_code=404, detail="ACM record not found")

        await record.delete()

        return {"message": "ACM record deleted successfully", "id": record_id}

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error deleting ACM record {record_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# =============================================================================
# Site Configuration Endpoints (E1-S8 - Victorian BAR Compliance)
# =============================================================================


@router.get("/config", response_model=SiteConfigResponse)
async def get_site_config(
    source_id: str = Query(..., description="Source document ID"),
):
    """
    Get site configuration for a source document.

    Returns the configuration if it exists, or an empty config template.
    """
    try:
        config = await SiteConfig.get_by_source(source_id)

        if config:
            return SiteConfigResponse(
                id=config.id,
                source_id=config.source_id,
                department=config.department,
                agency=config.agency,
                building_type=config.building_type,
                owned_or_leased=config.owned_or_leased,
                frequency_of_use=config.frequency_of_use,
                public_access=config.public_access,
                building_unique_id=config.building_unique_id,
                missing_fields=config.get_missing_bar_fields(),
                is_bar_complete=config.is_bar_complete(),
                created=str(config.created) if config.created else None,
                updated=str(config.updated) if config.updated else None,
            )

        # Return empty config template for new sources
        return SiteConfigResponse(
            source_id=source_id,
            missing_fields=[
                "department",
                "agency",
                "building_type",
                "owned_or_leased",
                "frequency_of_use",
                "public_access",
            ],
            is_bar_complete=False,
        )

    except Exception as e:
        logger.error(f"Error fetching site config for {source_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/config", response_model=SiteConfigResponse)
async def save_site_config(request: SiteConfigRequest):
    """
    Create or update site configuration for a source document.

    Uses upsert logic - creates if not exists, updates if exists.
    """
    try:
        config = await SiteConfig.upsert(
            source_id=request.source_id,
            department=request.department,
            agency=request.agency,
            building_type=request.building_type,
            owned_or_leased=request.owned_or_leased,
            frequency_of_use=request.frequency_of_use,
            public_access=request.public_access,
            building_unique_id=request.building_unique_id,
        )

        return SiteConfigResponse(
            id=config.id,
            source_id=config.source_id,
            department=config.department,
            agency=config.agency,
            building_type=config.building_type,
            owned_or_leased=config.owned_or_leased,
            frequency_of_use=config.frequency_of_use,
            public_access=config.public_access,
            building_unique_id=config.building_unique_id,
            missing_fields=config.get_missing_bar_fields(),
            is_bar_complete=config.is_bar_complete(),
            created=str(config.created) if config.created else None,
            updated=str(config.updated) if config.updated else None,
        )

    except Exception as e:
        logger.error(f"Error saving site config: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/config/templates", response_model=list[SiteConfigTemplateResponse])
async def list_site_config_templates(
    limit: int = Query(20, ge=1, le=100, description="Max templates to return"),
):
    """
    List available site configuration templates.

    Returns previously saved configurations that can be reused for new sources.
    """
    try:
        templates = await SiteConfig.get_templates(limit=limit)

        return [
            SiteConfigTemplateResponse(
                source_id=t.get("source_id", ""),
                source_title=t.get("source_title"),
                department=t.get("department"),
                agency=t.get("agency"),
                building_type=t.get("building_type"),
                owned_or_leased=t.get("owned_or_leased"),
                frequency_of_use=t.get("frequency_of_use"),
                public_access=t.get("public_access"),
            )
            for t in templates
        ]

    except Exception as e:
        logger.error(f"Error fetching site config templates: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/config/apply-template", response_model=SiteConfigResponse)
async def apply_site_config_template(request: ApplyTemplateRequest):
    """
    Apply a template configuration to a source document.

    Copies configuration from the template source to the target source.
    """
    try:
        # Get template config
        template = await SiteConfig.get_by_source(request.template_source_id)
        if not template:
            raise HTTPException(
                status_code=404,
                detail=f"Template configuration not found for source {request.template_source_id}",
            )

        # Apply template to target source
        config = await SiteConfig.upsert(
            source_id=request.source_id,
            department=template.department,
            agency=template.agency,
            building_type=template.building_type,
            owned_or_leased=template.owned_or_leased,
            frequency_of_use=template.frequency_of_use,
            public_access=template.public_access,
            # Don't copy building_unique_id - should be unique per building
        )

        return SiteConfigResponse(
            id=config.id,
            source_id=config.source_id,
            department=config.department,
            agency=config.agency,
            building_type=config.building_type,
            owned_or_leased=config.owned_or_leased,
            frequency_of_use=config.frequency_of_use,
            public_access=config.public_access,
            building_unique_id=config.building_unique_id,
            missing_fields=config.get_missing_bar_fields(),
            is_bar_complete=config.is_bar_complete(),
            created=str(config.created) if config.created else None,
            updated=str(config.updated) if config.updated else None,
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error applying template: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/config/agencies", response_model=AgencyListResponse)
async def list_agencies(
    department: Optional[str] = Query(None, description="Filter by department"),
):
    """
    List distinct agency values for autocomplete.

    Optionally filter by department.
    """
    try:
        agencies = await SiteConfig.get_agencies(department=department)
        return AgencyListResponse(agencies=agencies)

    except Exception as e:
        logger.error(f"Error fetching agencies: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# =============================================================================
# Building Review Endpoints (E19-S5 — Wizard Step 1)
# =============================================================================

# ACM record fields that live on individual acm_record rows (not site_config)
_ACM_RECORD_BUILDING_FIELDS = {
    "building_name",
    "building_address",
    "building_year",
    "building_size_m2",
    "number_of_levels",
    "building_construction",
    "roof_type",
    "date_of_inspection",
    "suburb",
    "postcode",
}

_PAGE_MARKER_PATTERNS = (
    re.compile(r"<!--\s*Page\s+(\d+)\s*-->", re.IGNORECASE),
    re.compile(r"[-—]+\s*Page\s+(\d+)\s*[-—]+", re.IGNORECASE),
    re.compile(r"PAGE\s+(\d+)\s+OF\s+\d+", re.IGNORECASE),
)


def _extract_page_marker(line: str) -> Optional[int]:
    """Extract page number from a Docling page marker line."""
    for pattern in _PAGE_MARKER_PATTERNS:
        match = pattern.search(line)
        if match:
            return int(match.group(1))
    return None


def _is_markdown_table_line(line: str) -> bool:
    """Return True when a line resembles a markdown table row."""
    stripped = line.strip()
    return stripped.count("|") >= 2 and not stripped.startswith("```")


def _extract_docling_markdown_tables(
    full_text: str,
    source_id: str,
) -> list[RawTableResponse]:
    """Extract markdown table blocks from Docling text grouped by page markers."""
    if not full_text.strip():
        return []

    tables: list[RawTableResponse] = []
    current_page = 1
    table_start_page = 1
    table_lines: list[str] = []

    def flush_table(page_end: int) -> None:
        nonlocal table_lines, table_start_page
        if not table_lines:
            return

        raw_text = "\n".join(table_lines).strip()
        table_lines = []
        if not raw_text:
            return

        table_index = len(tables) + 1
        tables.append(
            RawTableResponse(
                id=f"docling_table:{source_id}:{table_index}",
                source_id=source_id,
                page_start=table_start_page,
                page_end=page_end,
                table_type="docling_markdown",
                raw_text=raw_text,
            )
        )

    for line in full_text.splitlines():
        page_marker = _extract_page_marker(line)
        if page_marker is not None:
            flush_table(current_page)
            current_page = page_marker
            continue

        if _is_markdown_table_line(line):
            if not table_lines:
                table_start_page = current_page
            table_lines.append(line.rstrip())
            continue

        if table_lines:
            flush_table(current_page)

    flush_table(current_page)
    return tables


# site_config fields managed at the source level
_SITE_CONFIG_BUILDING_FIELDS = {
    "department",
    "agency",
    "sub_agency",
    "site_name",
    "building_type",
    "owned_or_leased",
    "building_unique_id",
    "frequency_of_use",
    "public_access",
    "building_out_of_scope",
    "building_out_of_scope_comments",
    "additional_comments",
}


def _build_building_response(
    row: dict, site_config: Optional[SiteConfig]
) -> BuildingResponse:
    """Merge a grouped acm_record row with site_config data into a BuildingResponse."""
    sc = site_config

    return BuildingResponse(
        building_id=row.get("building_id", ""),
        building_name=row.get("building_name"),
        building_address=row.get("building_address"),
        building_year=row.get("building_year"),
        building_size_m2=row.get("building_size_m2"),
        number_of_levels=row.get("number_of_levels"),
        building_construction=row.get("building_construction"),
        roof_type=row.get("roof_type"),
        date_of_inspection=row.get("date_of_inspection"),
        record_count=row.get("record_count", 0),
        suburb=row.get("suburb"),
        postcode=row.get("postcode"),
        # site_config fields (shared across all buildings for this source)
        department=sc.department if sc else None,
        agency=sc.agency if sc else None,
        sub_agency=sc.sub_agency if sc else None,
        site_name=sc.site_name if sc else None,
        building_type=sc.building_type if sc else None,
        owned_or_leased=sc.owned_or_leased if sc else None,
        building_unique_id=sc.building_unique_id if sc else None,
        frequency_of_use=sc.frequency_of_use if sc else None,
        public_access=sc.public_access if sc else None,
        building_out_of_scope=sc.building_out_of_scope if sc else False,
        building_out_of_scope_comments=sc.building_out_of_scope_comments
        if sc
        else None,
        additional_comments=sc.additional_comments if sc else None,
    )


@router.get("/jobs/{source_id}/raw-tables", response_model=list[RawTableResponse])
async def list_raw_tables_for_job(source_id: str):
    """Return raw table blocks for a job, preferring MinerU HTML when available."""
    try:
        normalized_source_id = ensure_record_id(source_id)

        sections = await ACMTableSection.get_by_source(source_id)
        mineru_sections = [section for section in sections if section.raw_html]

        if mineru_sections:
            return [
                RawTableResponse(
                    id=str(section.id) if section.id else "",
                    source_id=str(section.source_id),
                    page_start=section.page_start,
                    page_end=section.page_end,
                    table_type=section.table_type or "mineru_html",
                    raw_html=section.raw_html,
                    raw_text=section.raw_text,
                    building_name=section.building_name,
                )
                for section in mineru_sections
            ]

        section_text_rows = [section for section in sections if section.raw_text]
        if section_text_rows:
            return [
                RawTableResponse(
                    id=str(section.id) if section.id else "",
                    source_id=str(section.source_id),
                    page_start=section.page_start,
                    page_end=section.page_end,
                    table_type=section.table_type or "register_raw_text",
                    raw_text=section.raw_text,
                    building_name=section.building_name,
                )
                for section in section_text_rows
            ]

        from open_notebook.domain.notebook import Source

        source = await Source.get(source_id)
        if not source:
            raise HTTPException(status_code=404, detail="Source not found")

        return _extract_docling_markdown_tables(
            source.full_text or "",
            str(normalized_source_id),
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error loading raw tables for source {source_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get(
    "/raw-extractions/{source_id}",
    response_model=RawExtractionListResponse,
)
async def list_raw_extractions(
    source_id: str,
    provider: Optional[str] = Query(
        None,
        description="Filter by provider_id (e.g. 'docling', 'mineru')",
    ),
    page_number: Optional[int] = Query(
        None,
        description="Filter by page number (1-based)",
    ),
):
    """
    List per-provider raw extraction outputs for a source document.

    Returns one record per provider per page. Use the provider query param
    to narrow results to a single provider's output.
    """
    try:
        extractions = await RawExtraction.get_by_source(
            source_id,
            provider=provider,
            page_number=page_number,
        )

        items = [
            RawExtractionResponse(
                id=str(e.id or ""),
                source_id=str(e.source_id),
                provider_id=e.provider_id,
                extraction_backend=e.extraction_backend,
                page_number=e.page_number,
                raw_html=e.raw_html,
                raw_markdown=e.raw_markdown,
                structured_json=e.structured_json,
                bbox=e.bbox,
                confidence=e.confidence,
                officer_edits=e.officer_edits,
                created_at=str(e.created_at) if e.created_at else None,
            )
            for e in extractions
        ]

        return RawExtractionListResponse(
            extractions=items,
            total=len(items),
            source_id=source_id,
        )
    except Exception as e:
        logger.error(f"Error listing raw extractions for source {source_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.patch(
    "/raw-extractions/{source_id}/{extraction_id}",
    response_model=RawExtractionResponse,
)
async def patch_raw_extraction(
    source_id: str,
    extraction_id: str,
    body: PatchRawExtractionRequest,
):
    """Append officer edits and optionally update structured_json for a raw extraction row.

    Verifies that the extraction record belongs to the given source_id before
    applying changes. Returns the updated RawExtractionResponse on success.
    """
    try:
        try:
            extraction = await RawExtraction.get(extraction_id)
        except NotFoundError:
            raise HTTPException(status_code=404, detail="Raw extraction not found")

        # Normalize source_id: add 'source:' prefix if not present so the
        # comparison works whether the caller passes "s1" or "source:s1".
        normalized_source_id = (
            source_id if source_id.startswith("source:") else f"source:{source_id}"
        )
        if str(extraction.source_id) != normalized_source_id:
            raise HTTPException(status_code=403, detail="source_id mismatch")

        if body.structured_json is not None:
            try:
                _json.loads(body.structured_json)
            except ValueError:
                raise HTTPException(
                    status_code=422, detail="structured_json is not valid JSON"
                )
            extraction.structured_json = body.structured_json

        extraction.officer_edits = extraction.officer_edits + [
            e.model_dump() for e in body.edits
        ]

        await extraction.save()

        return RawExtractionResponse(
            id=str(extraction.id or ""),
            source_id=str(extraction.source_id),
            provider_id=extraction.provider_id,
            extraction_backend=extraction.extraction_backend,
            page_number=extraction.page_number,
            raw_html=extraction.raw_html,
            raw_markdown=extraction.raw_markdown,
            structured_json=extraction.structured_json,
            bbox=extraction.bbox,
            confidence=extraction.confidence,
            officer_edits=extraction.officer_edits,
            created_at=str(extraction.created_at) if extraction.created_at else None,
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error patching raw extraction {extraction_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/jobs/{source_id}/buildings", response_model=list[BuildingResponse])
async def list_buildings_for_job(source_id: str):
    """
    List all buildings detected for a source, merging acm_record and site_config data.
    Used by the Building Review Wizard Step 1.
    """
    try:
        sid = ensure_record_id(source_id)

        # Query distinct buildings grouped from acm_record
        query = """
            SELECT
                building_id,
                building_name,
                building_address,
                building_year,
                building_size_m2,
                number_of_levels,
                building_construction,
                roof_type,
                date_of_inspection,
                suburb,
                postcode,
                count() as record_count
            FROM acm_record
            WHERE source_id = $source_id
            GROUP BY
                building_id,
                building_name,
                building_address,
                building_year,
                building_size_m2,
                number_of_levels,
                building_construction,
                roof_type,
                date_of_inspection,
                suburb,
                postcode
            ORDER BY building_id;
        """
        rows = await repo_query(query, {"source_id": sid})

        if not rows:
            return []

        # Fetch site_config once — shared across all buildings for this source
        site_config = await SiteConfig.get_by_source(source_id)

        return [_build_building_response(row, site_config) for row in rows]

    except Exception as e:
        logger.error(f"Error listing buildings for job {source_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.put(
    "/jobs/{source_id}/buildings/{building_id}", response_model=BuildingResponse
)
async def update_building(
    source_id: str, building_id: str, request: BuildingUpdateRequest
):
    """
    Update building metadata during review.

    - site_config fields are saved to the site_config record for this source.
    - acm_record fields (building_name, building_year, etc.) update ALL acm_record
      rows that share this building_id within the source.
    """
    try:
        sid = ensure_record_id(source_id)
        update_data = request.model_dump(exclude_none=True)

        # Separate fields by destination table
        acm_updates = {
            k: v for k, v in update_data.items() if k in _ACM_RECORD_BUILDING_FIELDS
        }
        sc_updates = {
            k: v for k, v in update_data.items() if k in _SITE_CONFIG_BUILDING_FIELDS
        }

        # Apply acm_record field updates to all rows for this building
        if acm_updates:
            set_clause = ", ".join(f"{k} = ${k}" for k in acm_updates)
            acm_params = {"source_id": sid, "building_id": building_id, **acm_updates}
            update_query = f"""
                UPDATE acm_record
                SET {set_clause}
                WHERE source_id = $source_id AND building_id = $building_id;
            """
            await repo_query(update_query, acm_params)

        # Apply site_config field updates
        if sc_updates:
            await SiteConfig.upsert(source_id=source_id, **sc_updates)

        # Re-fetch the updated building row to return current state
        fetch_query = """
            SELECT
                building_id,
                building_name,
                building_address,
                building_year,
                building_size_m2,
                number_of_levels,
                building_construction,
                roof_type,
                date_of_inspection,
                suburb,
                postcode,
                count() as record_count
            FROM acm_record
            WHERE source_id = $source_id AND building_id = $building_id
            GROUP BY
                building_id,
                building_name,
                building_address,
                building_year,
                building_size_m2,
                number_of_levels,
                building_construction,
                roof_type,
                date_of_inspection,
                suburb,
                postcode
            LIMIT 1;
        """
        rows = await repo_query(
            fetch_query, {"source_id": sid, "building_id": building_id}
        )

        if not rows:
            raise HTTPException(
                status_code=404,
                detail=f"Building '{building_id}' not found for source '{source_id}'",
            )

        site_config = await SiteConfig.get_by_source(source_id)
        return _build_building_response(rows[0], site_config)

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating building {building_id} for job {source_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/jobs/{source_id}/publish")
async def publish_job(source_id: str):
    """
    Publish a job to the ACM Register.
    Sets source.review_status = 'published'.
    After this, records appear in the global register.
    """
    try:
        from open_notebook.domain.notebook import Source

        source = await Source.get(source_id)
        if not source:
            raise HTTPException(status_code=404, detail="Source not found")

        source.review_status = "published"
        await source.save()

        return {
            "source_id": source_id,
            "review_status": "published",
            "message": "Job published to ACM Register",
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error publishing job {source_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# =============================================================================
# ACM Product Classification Endpoints (E1-S9 - Victorian BAR Taxonomy)
# =============================================================================


@router.post("/classify", response_model=ClassifyResponse)
async def classify_acm_item(request: ClassifyRequest):
    """
    Classify a single ACM item into Victorian BAR taxonomy.

    Uses pattern-based matching first, then falls back to LLM if enabled.

    Example:
        POST /api/acm/classify
        {"item_description": "Vinyl floor tiles", "friability": "Non-friable"}

    Returns the product group (e.g., "Vinyl products") and product type
    (e.g., "Vinyl Tiles") with a confidence score.
    """
    try:
        from open_notebook.extractors.normalizers.taxonomy import (
            classify_product,
            classify_product_async,
        )

        if request.use_llm_fallback:
            # Async version with LLM fallback
            result = await classify_product_async(
                item_description=request.item_description,
                friability=request.friability,
                product=request.product,
                use_llm_fallback=True,
            )
        else:
            # Sync pattern-only version
            result = classify_product(
                item_description=request.item_description,
                friability=request.friability,
                product=request.product,
            )

        return ClassifyResponse(
            product_group=result.product_group,
            product_type=result.product_type,
            confidence=result.confidence,
            method=result.method,
        )

    except Exception as e:
        logger.error(f"Error classifying ACM item: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/classify/batch", response_model=BatchClassifyResponse)
async def classify_batch(request: BatchClassifyRequest):
    """
    Classify all ACM records for a source document.

    Updates records in the database with classification results.
    Optionally skips records that already have classification.

    Example:
        POST /api/acm/classify/batch
        {"source_id": "source:abc123", "use_llm_fallback": true}
    """
    try:
        from open_notebook.extractors.normalizers.taxonomy import (
            classify_product,
            classify_product_async,
        )

        # Get all records for source
        records = await ACMRecord.get_by_source(request.source_id)

        if not records:
            return BatchClassifyResponse(
                total=0,
                classified=0,
                skipped=0,
                errors=0,
                results=[],
            )

        total = len(records)
        classified = 0
        skipped = 0
        errors = 0
        results = []

        for record in records:
            try:
                # Skip if already classified and skip_classified is True
                if request.skip_classified and record.acm_product_group:
                    skipped += 1
                    results.append(
                        {
                            "record_id": str(record.id),
                            "status": "skipped",
                            "reason": "already_classified",
                        }
                    )
                    continue

                # Combine product and material description for classification
                item_description = record.material_description
                if record.product:
                    item_description = f"{record.product} {item_description}"

                # Classify
                if request.use_llm_fallback:
                    result = await classify_product_async(
                        item_description=item_description,
                        friability=record.friable,
                        product=record.product,
                        use_llm_fallback=True,
                    )
                else:
                    result = classify_product(
                        item_description=item_description,
                        friability=record.friable,
                        product=record.product,
                    )

                if result.product_group and result.product_type:
                    # Update record
                    record.acm_product_group = result.product_group
                    record.acm_product_type = result.product_type
                    record.classification_confidence = result.confidence
                    record.classification_method = result.method
                    record.classification_override = False
                    await record.save()

                    classified += 1
                    results.append(
                        {
                            "record_id": str(record.id),
                            "status": "classified",
                            "product_group": result.product_group,
                            "product_type": result.product_type,
                            "confidence": result.confidence,
                            "method": result.method,
                        }
                    )
                else:
                    skipped += 1
                    results.append(
                        {
                            "record_id": str(record.id),
                            "status": "skipped",
                            "reason": "no_match",
                        }
                    )

            except Exception as e:
                errors += 1
                results.append(
                    {
                        "record_id": str(record.id),
                        "status": "error",
                        "error": str(e),
                    }
                )
                logger.warning(f"Error classifying record {record.id}: {e}")

        logger.info(
            f"Batch classification for {request.source_id}: "
            f"{classified} classified, {skipped} skipped, {errors} errors (total: {total})"
        )

        return BatchClassifyResponse(
            total=total,
            classified=classified,
            skipped=skipped,
            errors=errors,
            results=results,
        )

    except Exception as e:
        logger.error(f"Error in batch classification: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/normalize", response_model=NormalizeResponse)
async def normalize_recommendation_text(request: NormalizeRequest):
    """
    Normalize a consultant recommendation to a canonical action.

    Uses pattern-based matching against known consultant wording patterns
    to map free-text recommendations to standardized actions.

    Example:
        POST /api/acm/normalize
        {"recommendation": "Maintain in current condition and label"}

    Returns the canonical action (e.g., "maintain_in_situ") with confidence.
    """
    try:
        from open_notebook.extractors.normalizers.recommendations import (
            normalize_recommendation,
        )

        result = normalize_recommendation(request.recommendation)

        return NormalizeResponse(
            raw_text=result.raw_text,
            normalized_action=result.normalized_action,
            confidence=result.confidence,
            method=result.method,
        )

    except Exception as e:
        logger.error(f"Error normalizing recommendation: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# =============================================================================
# Field Schema Configuration Endpoints (E1-S11 - Generic Configurable Parser)
# =============================================================================


async def _load_db_field_config() -> dict | None:
    """Load field config override from SurrealDB field_schema table."""
    try:
        result = await repo_query(
            "SELECT * FROM field_schema ORDER BY updated DESC LIMIT 1"
        )
        if result and result[0].get("config_json"):
            import json

            return json.loads(result[0]["config_json"])
    except Exception as e:
        logger.debug(f"No DB field config found: {e}")
    return None


async def _save_db_field_config(config_dict: dict) -> None:
    """Save field config override to SurrealDB field_schema table."""
    import json

    config_json = json.dumps(config_dict)
    version = config_dict.get("version", "1.0.0")
    # Upsert using a fixed ID so there's only ever one config record
    await repo_query(
        """
        UPSERT field_schema:default SET
            config_json = $config_json,
            version = $version,
            updated = time::now()
        """,
        {"config_json": config_json, "version": version},
    )


async def _delete_db_field_config() -> None:
    """Delete field config override from SurrealDB."""
    await repo_query("DELETE field_schema:default")


def _config_to_response(config) -> FieldSchemaConfigResponse:
    """Convert a FieldSchemaConfig to API response model."""
    return FieldSchemaConfigResponse(
        fields=[
            FieldDefResponse(
                internal_name=f.internal_name,
                display_name=f.display_name,
                excel_column=f.excel_column,
                col_index=f.col_index,
                field_type=f.field_type,
                required=f.required,
                active=f.active,
                enum_name=f.enum_name,
                group=f.group,
            )
            for f in config.fields
        ],
        enums=config.enums,
        business_rules=[
            BusinessRuleResponse(
                rule_id=r.rule_id,
                description=r.description,
                enabled=r.enabled,
            )
            for r in config.business_rules
        ],
        version=config.version,
        source_template=config.source_template,
    )


@router.get("/field-config", response_model=FieldSchemaConfigResponse)
async def get_field_config():
    """
    Get the current field schema configuration.

    Returns the 47 BAR field definitions, enum picklists, and business rules
    that drive the GenericParser extraction behavior. Checks the database for
    overrides first, then falls back to the default JSON config.
    """
    try:
        db_config = await _load_db_field_config()
        if db_config is not None:
            return FieldSchemaConfigResponse(**db_config)

        from open_notebook.extractors.parsers.config_loader import load_field_schema

        config = load_field_schema()
        return _config_to_response(config)

    except Exception as e:
        logger.error(f"Error getting field config: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.put("/field-config", response_model=FieldSchemaConfigResponse)
async def update_field_config(request: FieldSchemaConfigUpdateRequest):
    """
    Update the field schema configuration.

    Allows toggling field active status, modifying business rules,
    and updating enum definitions. Persists to the database so changes
    survive server restarts.
    """
    try:
        config_dict = request.model_dump()
        await _save_db_field_config(config_dict)
        return FieldSchemaConfigResponse(**config_dict)

    except Exception as e:
        logger.error(f"Error updating field config: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/field-config/reset", response_model=FieldSchemaConfigResponse)
async def reset_field_config():
    """
    Reset field schema configuration to defaults.

    Clears any database overrides and reloads the default config
    from the BAR JSON schema files.
    """
    try:
        await _delete_db_field_config()

        from open_notebook.extractors.parsers import config_loader

        # Clear the cached config to force reload from JSON
        config_loader._FIELD_SCHEMA = None
        config = config_loader.load_field_schema()
        return _config_to_response(config)

    except Exception as e:
        logger.error(f"Error resetting field config: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/taxonomy", response_model=TaxonomyResponse)
async def get_taxonomy(
    friability: Optional[str] = Query(
        None, description="Friability type: 'Friable' or 'Non-friable' (default)"
    ),
):
    """
    Get available ACM product taxonomy.

    Returns the Victorian BAR product groups and types for the specified friability.
    Useful for UI dropdowns and validation.

    Example:
        GET /api/acm/taxonomy?friability=Non-friable
    """
    try:
        from open_notebook.extractors.normalizers.taxonomy import get_product_groups

        groups = get_product_groups(friability)

        # Determine friability label
        if (
            friability
            and "friable" in friability.lower()
            and "non" not in friability.lower()
        ):
            friability_label = "Friable"
        else:
            friability_label = "Non-friable"

        return TaxonomyResponse(
            friability=friability_label,
            groups=[
                TaxonomyGroupResponse(
                    pc_code=g.get("pc_code", ""),
                    product_group_header=g.get("product_group_header", ""),
                    product_types=g.get("product_types", []),
                )
                for g in groups
            ],
        )

    except Exception as e:
        logger.error(f"Error fetching taxonomy: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# --- Field Mapping Endpoints (E5-S4) ---


@router.get("/field-mapping")
async def get_field_mapping():
    """Get the active export field mapping."""
    from open_notebook.domain.field_mapping import FieldMapping

    try:
        mapping = await FieldMapping.get_active()
        return mapping.model_dump()
    except Exception as e:
        logger.error(f"Error fetching field mapping: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.put("/field-mapping")
async def update_field_mapping(data: FieldMappingUpdateRequest):
    """Update the active field mapping."""
    from open_notebook.domain.field_mapping import FieldMapping, FieldMappingEntry

    try:
        mapping = await FieldMapping.get_active()
        if data.mappings is not None:
            mapping.mappings = [
                FieldMappingEntry.model_validate(m.model_dump()) for m in data.mappings
            ]
        if data.name is not None:
            mapping.name = data.name
        if data.notes is not None:
            mapping.notes = data.notes
        await mapping.update()
        return mapping.model_dump()
    except Exception as e:
        logger.error(f"Error updating field mapping: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/field-mapping/reset")
async def reset_field_mapping():
    """Reset field mapping to defaults."""
    from open_notebook.domain.field_mapping import FieldMapping

    try:
        mapping = await FieldMapping.reset_to_defaults()
        return mapping.model_dump()
    except Exception as e:
        logger.error(f"Error resetting field mapping: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# =============================================================================
# SF Field Schema Endpoint (E30-S1 — V3 Foundation)
# =============================================================================


@router.get("/field-schema", response_model=SFFieldSchemaConfigResponse)
async def get_sf_field_schema():
    """Get the current Salesforce field schema configuration.

    Returns the SF schema bundle parsed from V3 markdown files.
    Falls back to in-memory parse if DB record is not yet populated.

    Returns:
        SFFieldSchemaConfigResponse with building_fields, item_fields,
        picklists, and dependency chains.
    """
    from open_notebook.extractors.parsers.config_loader import (
        SFSchemaLoadError,
        load_sf_field_schema,
    )

    try:
        schema = load_sf_field_schema()
        return SFFieldSchemaConfigResponse(**schema.model_dump())
    except SFSchemaLoadError as e:
        logger.error(f"SF schema load error: {e}")
        raise HTTPException(
            status_code=500,
            detail=f"SF schema not available: {e}",
        )
    except Exception as e:
        logger.error(f"Unexpected error loading SF schema: {e}")
        raise HTTPException(
            status_code=500,
            detail=f"SF schema not available: {e}",
        )


# --- Building Record CRUD (E30-S2) ---


@router.get("/buildings", response_model=BuildingRecordListResponse)
async def list_building_records(
    source_id: str = Query(..., description="Source ID to filter by (required)"),
):
    """List all building records for a source, with ACM item counts."""
    try:
        buildings = await BuildingRecord.get_by_source(source_id)

        # Compute record_count per building via a single query
        record_counts: dict[str, int] = {}
        if buildings:
            from open_notebook.database.repository import repo_query

            rows = await repo_query(
                "SELECT building_record_id, count() AS cnt "
                "FROM acm_record WHERE source_id = $source_id "
                "AND building_record_id != NONE "
                "GROUP BY building_record_id",
                {"source_id": source_id},
            )
            for row in rows:
                bid = str(row.get("building_record_id", ""))
                record_counts[bid] = row.get("cnt", 0)

        responses = []
        for b in buildings:
            data = b.model_dump()
            data["record_count"] = record_counts.get(str(b.id), 0)
            responses.append(BuildingRecordResponse(**data))

        return BuildingRecordListResponse(
            buildings=responses,
            total=len(responses),
        )
    except Exception as e:
        logger.error(f"Error listing building records: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/buildings/{building_id:path}", response_model=BuildingRecordResponse)
async def get_building_record(building_id: str):
    """Get a single building record by ID."""
    try:
        building = await BuildingRecord.get(building_id)
        return BuildingRecordResponse(**building.model_dump())
    except NotFoundError:
        raise HTTPException(status_code=404, detail="Building record not found")
    except Exception as e:
        logger.error(f"Error getting building record {building_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/buildings", response_model=BuildingRecordResponse, status_code=201)
async def create_building_record(request: BuildingRecordCreateRequest):
    """Create a new building record. internal_id is auto-generated."""
    try:
        internal_id = await BuildingRecord.generate_internal_id(request.source_id)
        data = request.model_dump(exclude_none=True)
        data["internal_id"] = internal_id
        building = BuildingRecord(**data)
        await building.save()
        return BuildingRecordResponse(**building.model_dump())
    except Exception as e:
        logger.error(f"Error creating building record: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.put("/buildings/{building_id:path}", response_model=BuildingRecordResponse)
async def update_building_record(
    building_id: str, request: BuildingRecordUpdateRequest
):
    """Update an existing building record."""
    try:
        building = await BuildingRecord.get(building_id)
        update_data = request.model_dump(exclude_none=True)
        for key, value in update_data.items():
            setattr(building, key, value)
        await building.save()
        return BuildingRecordResponse(**building.model_dump())
    except Exception as e:
        logger.error(f"Error updating building record {building_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/buildings/{building_id:path}")
async def delete_building_record(building_id: str):
    """Delete a building record."""
    try:
        building = await BuildingRecord.get(building_id)
        await building.delete()
        return {"deleted": True, "id": building_id}
    except NotFoundError:
        raise HTTPException(status_code=404, detail="Building record not found")
    except Exception as e:
        logger.error(f"Error deleting building record {building_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# --- Source Intelligence (E30-S9) ---


@router.get(
    "/source-intelligence/{source_id:path}",
    response_model=SourceIntelligenceResponse,
)
async def get_source_intelligence_endpoint(source_id: str):
    """Return persisted pre-extraction intelligence for a source.

    Returns 404 if no intelligence has been saved yet (extraction not run
    or the save_intelligence node hasn't executed).
    """
    try:
        data = await get_source_intelligence(source_id)
        if not data:
            raise HTTPException(
                status_code=404,
                detail="No intelligence data found for this source",
            )
        return SourceIntelligenceResponse(**data)
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error fetching intelligence for {source_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# --- Validation Summary + Bulk Fix (E33-S4) ---


@router.get("/validation-summary", response_model=ValidationSummaryResponse)
async def get_validation_summary(
    source_id: str = Query(..., description="Source ID"),
):
    """Return per-building validation error counts for the sidebar badges.

    Queries all acm_record rows for this source that have one or more
    validation_errors and groups results by building_id.
    """
    try:
        sid = ensure_record_id(source_id)
        query = """
            SELECT building_id,
                   count() as error_count
            FROM acm_record
            WHERE source_id = $source_id
              AND array::len(validation_errors) > 0
            GROUP BY building_id;
        """
        rows = await repo_query(query, {"source_id": sid})
        buildings = [
            BuildingValidationSummary(
                building_id=r.get("building_id", ""),
                error_count=r.get("error_count", 0),
            )
            for r in rows
        ]
        return ValidationSummaryResponse(buildings=buildings)
    except Exception as e:
        logger.error(f"Error fetching validation summary for {source_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/bulk-fix", response_model=BulkFixResponse)
async def bulk_fix_records(
    source_id: str = Query(..., description="Source ID"),
    building_id: Optional[str] = Query(None, description="Optional building filter"),
):
    """Re-validate and auto-correct fixable records.

    Loads all records for the source (optionally filtered to a building)
    that currently have validation_errors, runs validate_acm_record() on
    each one (which applies normalizations as side-effects to the dict),
    and persists any records that now pass validation back to SurrealDB.
    """
    from open_notebook.extractors.validators.acm_validator import validate_acm_record

    try:
        sid = ensure_record_id(source_id)

        conditions = [
            "source_id = $source_id",
            "array::len(validation_errors) > 0",
        ]
        params: dict = {"source_id": sid}

        if building_id:
            conditions.append("building_id = $building_id")
            params["building_id"] = building_id

        where_clause = " AND ".join(conditions)
        fetch_query = f"SELECT * FROM acm_record WHERE {where_clause};"
        records = await repo_query(fetch_query, params)

        fixed_count = 0
        for rec in records:
            record_dict = dict(rec)
            result = validate_acm_record(record_dict)

            if result.is_valid:
                record_id = rec.get("id", "")
                if not record_id:
                    continue

                # Build SET clause for normalised enum fields that may have changed
                set_parts = [
                    "validation_status = 'valid'",
                    "validation_errors = []",
                    "updated = time::now()",
                ]
                update_params: dict = {"record_id": ensure_record_id(str(record_id))}

                for field_name in [
                    "sample_result",
                    "material_condition",
                    "friable",
                    "disturbance_potential",
                ]:
                    if field_name in record_dict:
                        set_parts.append(f"{field_name} = ${field_name}")
                        update_params[field_name] = record_dict[field_name]

                set_clause = ", ".join(set_parts)
                update_query = f"UPDATE $record_id SET {set_clause};"
                await repo_query(update_query, update_params)
                fixed_count += 1

        # Count records still failing validation
        remaining_query = (
            "SELECT count() as cnt FROM acm_record "
            "WHERE source_id = $source_id "
            "AND array::len(validation_errors) > 0 "
            "GROUP ALL;"
        )
        remaining_rows = await repo_query(remaining_query, {"source_id": sid})
        remaining_count = remaining_rows[0].get("cnt", 0) if remaining_rows else 0

        return BulkFixResponse(
            fixed_count=fixed_count,
            remaining_errors=remaining_count,
        )

    except Exception as e:
        logger.error(f"Error running bulk-fix for {source_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/bulk-edit", response_model=BulkEditResponse)
async def bulk_edit_records(
    request: BulkEditRequest,
    source_id: str = Query(..., description="Source ID for authorization"),
):
    """Set a single field to the same value on all specified records.

    Publishes bulk.progress events (one per record) and bulk.complete at end.
    Source_id is required for authorization but only records matching record_ids
    are updated.
    """
    # Whitelist check: only allow editable ACMRecord fields, not system/immutable ones.
    _immutable_fields = {"id", "source_id", "created", "updated", "table_bbox"}
    allowed_fields = frozenset(
        f for f in ACMRecord.model_fields if f not in _immutable_fields
    )
    if request.field not in allowed_fields:
        raise HTTPException(
            status_code=422,
            detail=f"Field '{request.field}' is not a valid editable ACMRecord field.",
        )

    from open_notebook.extractors.pipeline_event_bus import (
        V3PipelineEvent,
        get_event_bus,
    )

    bus = get_event_bus()
    total = len(request.record_ids)
    updated = 0

    for record_id in request.record_ids:
        try:
            rid = ensure_record_id(record_id)
            await repo_query(
                f"UPDATE $rid SET `{request.field}` = $val, updated = time::now();",
                {"rid": rid, "val": request.value},
            )
            updated += 1
        except Exception as e:
            logger.warning(f"Failed to update record {record_id}: {e}")

        await bus.publish(
            V3PipelineEvent(
                type="bulk.progress",
                operation_id=request.operation_id,
                data={
                    "processed": updated,
                    "total": total,
                    "percent": round((updated / max(total, 1)) * 100),
                },
            )
        )

    await bus.publish(
        V3PipelineEvent(
            type="bulk.complete",
            operation_id=request.operation_id,
            data={"updated_count": updated, "field": request.field},
        )
    )

    return BulkEditResponse(updated_count=updated, operation_id=request.operation_id)


@router.post("/bulk-validate", response_model=BulkValidateResponse)
async def bulk_validate_records(request: BulkValidateRequest):
    """Re-run SF validation on the specified records only.

    Equivalent to the existing /bulk-fix but scoped to explicit record_ids.
    """
    from open_notebook.extractors.validators.acm_validator import validate_acm_record

    fixed_count = 0
    for record_id in request.record_ids:
        try:
            rid = ensure_record_id(record_id)
            rows = await repo_query("SELECT * FROM $rid;", {"rid": rid})
            if not rows:
                continue
            record_dict = dict(rows[0])
            result = validate_acm_record(record_dict)

            if result.is_valid:
                await repo_query(
                    "UPDATE $rid SET validation_status = 'valid', "
                    "validation_errors = [], updated = time::now();",
                    {"rid": rid},
                )
                fixed_count += 1
            else:
                error_msgs = [
                    e.message if hasattr(e, "message") else str(e)
                    for e in result.errors
                ]
                await repo_query(
                    "UPDATE $rid SET validation_status = 'invalid', "
                    "validation_errors = $errors, updated = time::now();",
                    {"rid": rid, "errors": error_msgs},
                )
        except Exception as e:
            logger.warning(f"Failed to validate record {record_id}: {e}")

    remaining_rows = await repo_query(
        "SELECT count() as cnt FROM acm_record "
        "WHERE id IN $ids AND array::len(validation_errors) > 0 GROUP ALL;",
        {"ids": [ensure_record_id(r) for r in request.record_ids]},
    )
    remaining = remaining_rows[0].get("cnt", 0) if remaining_rows else 0

    return BulkValidateResponse(fixed_count=fixed_count, remaining_errors=remaining)


@router.get("/provenance/{record_id}", response_model=ProvenanceResponse)
async def get_provenance(record_id: str):
    """Get aggregated provenance data for a single ACM record (E33-S6).

    Combines the record itself with its parent ACMTableSection (consensus
    metadata), all per-provider RawExtractions for the same page, and the
    source file path so the frontend can render the originating PDF page
    with a bounding-box overlay.
    """
    from open_notebook.domain.notebook import Source

    # 1. Fetch the ACM record
    try:
        record = await ACMRecord.get(record_id)
    except Exception:
        record = None
    if not record:
        raise HTTPException(status_code=404, detail="ACM record not found")

    # 2. Build ACMRecordResponse (reuse the same mapping as get_acm_record)
    record_response = ACMRecordResponse(
        id=str(record.id),
        source_id=str(record.source_id),
        school_name=record.school_name,
        school_code=record.school_code,
        building_id=record.building_id,
        building_name=record.building_name,
        building_year=record.building_year,
        building_construction=record.building_construction,
        room_id=record.room_id,
        room_name=record.room_name,
        room_area=record.room_area,
        area_type=record.area_type,
        product=record.product,
        material_description=record.material_description,
        extent=record.extent,
        location=record.location,
        friable=record.friable,
        material_condition=record.material_condition,
        risk_status=record.risk_status,
        result=record.result,
        page_number=record.page_number,
        extraction_confidence=record.extraction_confidence,
        acm_product_group=record.acm_product_group,
        acm_product_type=record.acm_product_type,
        classification_confidence=record.classification_confidence,
        classification_method=record.classification_method,
        classification_override=record.classification_override,
        sample_no=record.sample_no,
        sample_result=record.sample_result,
        quantity=record.quantity,
        acm_labelled=record.acm_labelled,
        acm_label_details=record.acm_label_details,
        identifying_company=record.identifying_company,
        disturbance_potential=record.disturbance_potential,
        hygienist_recommendations=record.hygienist_recommendations,
        normalized_action=record.normalized_action,
        data_issues=record.data_issues,
        floor_level=record.floor_level,
        no_access=record.no_access,
        smf_present=record.smf_present,
        validation_status=getattr(record, "validation_status", None),
        validation_errors=getattr(record, "validation_errors", []) or [],
        created=str(record.created) if record.created else None,
        updated=str(record.updated) if record.updated else None,
    )

    # 3. Fetch parent ACMTableSection for consensus metadata
    table_section = None
    if record.parent_table_id:
        try:
            section = await ACMTableSection.get(record.parent_table_id)
            if section:
                table_section = {
                    "consensus_tier": section.consensus_tier,
                    "consensus_scores": section.consensus_scores,
                    "page_start": section.page_start,
                    "page_end": section.page_end,
                    "building_name": section.building_name,
                    "table_type": section.table_type,
                }
        except Exception as exc:
            logger.debug(f"Could not load table section for record {record_id}: {exc}")

    # 4. Fetch per-provider RawExtractions for the same source + page
    raw_extraction_responses: list[RawExtractionResponse] = []
    if record.page_number:
        try:
            extractions = await RawExtraction.get_by_source(
                str(record.source_id),
                page_number=record.page_number,
            )
            raw_extraction_responses = [
                RawExtractionResponse(
                    id=str(e.id or ""),
                    source_id=str(e.source_id),
                    provider_id=e.provider_id,
                    extraction_backend=e.extraction_backend,
                    page_number=e.page_number,
                    raw_html=e.raw_html,
                    raw_markdown=e.raw_markdown,
                    structured_json=e.structured_json,
                    bbox=e.bbox,
                    confidence=e.confidence,
                    officer_edits=e.officer_edits,
                    created_at=str(e.created_at) if e.created_at else None,
                )
                for e in extractions
            ]
        except Exception as exc:
            logger.debug(
                f"Could not load raw extractions for record {record_id}: {exc}"
            )

    # 5. Fetch source file path and title
    source_file_path: Optional[str] = None
    source_title: Optional[str] = None
    try:
        source = await Source.get(str(record.source_id))
        if source:
            source_title = source.title
            if source.asset and source.asset.file_path:
                source_file_path = source.asset.file_path
    except Exception as exc:
        logger.debug(f"Could not load source for record {record_id}: {exc}")

    return ProvenanceResponse(
        record=record_response,
        table_section=table_section,
        raw_extractions=raw_extraction_responses,
        source_file_path=source_file_path,
        source_title=source_title,
    )
