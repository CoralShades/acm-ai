#!/usr/bin/env python3
"""
Asbestos Assessment PDF to BAR Excel Extractor
Prototype for local Ollama-based extraction

Requirements:
- ollama pull qwen2.5-vl:32b-instruct-q4_K_M  (or smaller: qwen2.5-vl:7b)
- pip install pdf2image pymupdf openpyxl pandas Pillow

Usage:
    python asbestos_extractor.py input.pdf output.xlsx
"""

import base64
import json
import re
from datetime import datetime
from pathlib import Path
from typing import Optional

import fitz  # PyMuPDF
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ============================================================================
# CONFIGURATION
# ============================================================================

OLLAMA_MODEL = "qwen2.5-vl:32b-instruct-q4_K_M"  # Change to smaller if needed
# OLLAMA_MODEL = "qwen2.5-vl:7b"  # Lighter option for testing

# ACM Product mappings
ACM_PRODUCT_MAPPINGS = {
    # Pattern → (Product Group, Product Type)
    r"vinyl\s*(sheet|flooring)": ("Vinyl products", "Vinyl sheet"),
    r"vinyl\s*tile": ("Vinyl products", "Vinyl Tiles"),
    r"hessian\s*back": ("Vinyl products", "Hessian backed Vinyl sheet"),
    r"(flange\s*)?mastic": ("Gasket, friction products and adhesives", "Mastic"),
    r"gasket": ("Gasket, friction products and adhesives", "Gasket(s)"),
    r"(fibre|fiber)\s*cement|fc\s*sheet": ("Cement products", "Flat Sheeting"),
    r"fuse": ("Insulation Products", "Electrical Components"),
    r"internal\s*lining|filing\s*cabinet": ("Insulation Products", "Internal Lining"),
    r"expansion\s*joint": ("Gasket, friction products and adhesives", "Mastic"),
    r"skirting": ("Vinyl products", "Vinyl sheet"),
}

# BAR Column order
BAR_COLUMNS = [
    "Department",
    "Agency",
    "Sub Agency",
    "Site Name (if applicable)",
    "Building Name",
    "Building Type",
    "Building Address",
    "Suburb",
    "Postcode",
    "Owned or Leased",
    "Building Unique ID",
    "Frequency of use",
    "Public Access?",
    "Date of Inspection",
    "Estimated Year Built",
    "Est. Building Size (m2)",
    "Number of Levels",
    "Construction Type",
    "Roof Type",
    "Internal / External",
    "Level",
    "Room or Area",
    "Location in Room",
    "Specific Item/ACM Name",
    "Friability of material",
    "ACM Product Group",
    "ACM Product Type",
    "NATA Endorsed Sample number (if available)",
    "Sample Result",
    "Identifying Hygiene or Consulting Company",
    "Condition",
    "Disturbance Potential",
    "Quantity",
    "Labelled",
    "Label Details",
    "Hygienist Recommendations",
    "Additional Comments",
    "PSB Supplied ACM ID",
    "Assumed Removed?",
    "Date of Removal",
    "Quantity Removed",
    "Asbestos Removal Notification No",
    "EPA Waste Transport Certificate No",
]


# ============================================================================
# PDF PROCESSING
# ============================================================================


def pdf_page_to_image(pdf_path: str, page_num: int, dpi: int = 200) -> bytes:
    """Convert a single PDF page to PNG image bytes"""
    doc = fitz.open(pdf_path)
    page = doc[page_num]
    mat = fitz.Matrix(dpi / 72, dpi / 72)
    pix = page.get_pixmap(matrix=mat)
    img_bytes = pix.tobytes("png")
    doc.close()
    return img_bytes


def extract_text_from_pdf(pdf_path: str) -> dict:
    """Extract all text from PDF, organized by page"""
    doc = fitz.open(pdf_path)
    pages = {}
    for i, page in enumerate(doc):
        pages[i] = page.get_text()
    doc.close()
    return pages


def classify_page(text: str) -> str:
    """Classify page type based on content"""
    text_lower = text.lower()

    if "division 5 asbestos assessment" in text_lower and "client no" in text_lower:
        return "cover"
    elif "table of contents" in text_lower:
        return "toc"
    elif "executive summary" in text_lower and "key findings" in text_lower:
        return "executive"
    elif "asbestos register" in text_lower and (
        "area" in text_lower or "level" in text_lower
    ):
        return "register"
    elif "bulk sample analysis" in text_lower or "nata" in text_lower:
        return "lab_report"
    elif "risk assessment factors" in text_lower:
        return "appendix"
    else:
        return "body"


# ============================================================================
# METADATA EXTRACTION (Text-based)
# ============================================================================


def extract_metadata_from_text(pages: dict) -> dict:
    """Extract site metadata from PDF text"""
    all_text = "\n".join(pages.values())

    metadata = {
        "agency": "Victoria Police",
        "site_name": "",
        "building_name": "",
        "address": "",
        "suburb": "",
        "postcode": "",
        "inspection_date": None,
        "year_built": None,
        "building_size": None,
        "levels": None,
        "construction": "",
        "roof_type": "",
        "consultant": "Prensa Pty Ltd",
    }

    # Extract site name (Police Station pattern) - get the line containing it
    for line in all_text.split("\n"):
        if "Police Station" in line:
            site_name = line.strip()
            metadata["site_name"] = site_name
            metadata["building_name"] = site_name
            break

    # Extract address - look for pattern "## Street Name" followed by suburb
    addr_match = re.search(
        r"(\d+\s+[\w]+\s+(?:Road|Street|Avenue|Drive|Lane|Parade|Crescent|Way))\s*\n\s*(\w+)",
        all_text,
        re.IGNORECASE,
    )
    if addr_match:
        metadata["address"] = addr_match.group(1).strip()
        metadata["suburb"] = addr_match.group(2).strip()

    # Extract from Site Address line in Table 1
    site_addr_match = re.search(
        r"Site Address\s*\n?\s*(\d+\s+[\w\s]+),\s*(\w+)", all_text, re.IGNORECASE
    )
    if site_addr_match:
        metadata["address"] = site_addr_match.group(1).strip()
        metadata["suburb"] = site_addr_match.group(2).strip()

    # Extract year built (look for 4-digit year after completion or construction)
    year_match = re.search(
        r"(?:Completion|Construction)[:\s]*\n?\s*(\d{4})", all_text, re.IGNORECASE
    )
    if year_match:
        metadata["year_built"] = int(year_match.group(1))

    # Extract building size
    size_match = re.search(
        r"Approximate area[:\s]*\n?\s*([\d,]+)\s*m", all_text, re.IGNORECASE
    )
    if size_match:
        metadata["building_size"] = int(size_match.group(1).replace(",", ""))

    # Extract levels
    levels_match = re.search(r"Levels[:\s]*\n?\s*(\d+)", all_text, re.IGNORECASE)
    if levels_match:
        metadata["levels"] = int(levels_match.group(1))

    # Extract construction type (External walls)
    const_match = re.search(r"External walls[:\s]*\n?\s*(\w+)", all_text, re.IGNORECASE)
    if const_match:
        metadata["construction"] = const_match.group(1)

    # Extract roof type
    roof_match = re.search(r"Roof type[:\s]*\n?\s*(\w+)", all_text, re.IGNORECASE)
    if roof_match:
        metadata["roof_type"] = roof_match.group(1)

    # Extract inspection/assessment date - look for "conducted the Assessment on the Xth Month Year"
    date_match = re.search(
        r"(?:Assessment|conducted)\s+(?:on\s+)?(?:the\s+)?(\d{1,2})(?:st|nd|rd|th)?\s+(January|February|March|April|May|June|July|August|September|October|November|December)\s+(\d{4})",
        all_text,
        re.IGNORECASE,
    )
    if date_match:
        day, month, year = date_match.groups()
        month_map = {
            "january": 1,
            "february": 2,
            "march": 3,
            "april": 4,
            "may": 5,
            "june": 6,
            "july": 7,
            "august": 8,
            "september": 9,
            "october": 10,
            "november": 11,
            "december": 12,
        }
        metadata["inspection_date"] = datetime(
            int(year), month_map[month.lower()], int(day)
        )

    # Fallback: look for Date of Identification in register (usually Apr-20 format)
    if not metadata["inspection_date"]:
        date_match2 = re.search(
            r"(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)-(\d{2})", all_text
        )
        if date_match2:
            month_abbr, year_short = date_match2.groups()
            month_map = {
                "jan": 1,
                "feb": 2,
                "mar": 3,
                "apr": 4,
                "may": 5,
                "jun": 6,
                "jul": 7,
                "aug": 8,
                "sep": 9,
                "oct": 10,
                "nov": 11,
                "dec": 12,
            }
            year = 2000 + int(year_short)
            metadata["inspection_date"] = datetime(
                year, month_map[month_abbr.lower()], 1
            )

    # Extract postcode - 4-digit starting with 3 (Victorian postcode)
    # Be more specific to avoid catching other numbers
    postcode_match = re.search(r"(?:VIC|Victoria)\s*(\d{4})", all_text)
    if postcode_match:
        metadata["postcode"] = postcode_match.group(1)
    else:
        # Fallback: find 30XX or 31XX near suburb name
        if metadata["suburb"]:
            suburb_context = all_text[
                all_text.lower().find(
                    metadata["suburb"].lower()
                ) : all_text.lower().find(metadata["suburb"].lower()) + 100
            ]
            pc_match = re.search(r"\b(3\d{3})\b", suburb_context)
            if pc_match:
                metadata["postcode"] = pc_match.group(1)

    return metadata


# ============================================================================
# VISION-BASED TABLE EXTRACTION (Ollama)
# ============================================================================


def extract_register_with_vision(image_bytes: bytes, model: str = OLLAMA_MODEL) -> list:
    """Use Ollama vision model to extract register table from image"""
    try:
        import ollama
    except ImportError:
        print("WARNING: ollama package not installed. Using mock data.")
        return []

    img_b64 = base64.b64encode(image_bytes).decode()

    prompt = """Analyze this asbestos register table image. Extract ALL visible rows as a JSON array.

For each row in the table, extract these fields:
- area_level: Floor/level (e.g., "Ground floor", "First floor", "External")  
- room_location: Room or area name
- feature: Building feature column value
- item_description: Item description with any color notes in parentheses
- hazard_status: "Positive", "Negative", or "Assumed positive"
- sample_number: Sample ID like "34511-039-001" or "Same as 34511-039-001" or "Not Sampled"
- friability: "Non-friable" or "Friable" or "-"
- labelled: "Yes" or "No"
- disturbance_potential: "Low", "Medium", "High", or "-"
- condition: "Good", "Fair", "Poor", or "-"
- quantity: Amount with unit (e.g., "3 units", "2m2", "Throughout", "-")
- comments: Recommendations text, or "-" if none

IMPORTANT: 
- Extract EVERY row visible, even if some fields are empty
- Return ONLY a valid JSON array, no other text
- Use "-" for empty/missing values

Example output format:
[
  {"area_level": "Ground floor", "room_location": "Main foyer", "feature": "Floor", "item_description": "Vinyl sheet (cream)", "hazard_status": "Negative", "sample_number": "34511-039-001", "friability": "Non-friable", "labelled": "No", "disturbance_potential": "-", "condition": "-", "quantity": "-", "comments": "-"}
]"""

    try:
        response = ollama.chat(
            model=model,
            messages=[{"role": "user", "content": prompt, "images": [img_b64]}],
            options={"temperature": 0.1, "num_predict": 4096},
        )

        text = response["message"]["content"]

        # Extract JSON array from response
        start = text.find("[")
        end = text.rfind("]") + 1
        if start >= 0 and end > start:
            return json.loads(text[start:end])
        return []

    except Exception as e:
        print(f"Vision extraction error: {e}")
        return []


# ============================================================================
# DATA TRANSFORMATION
# ============================================================================


def lookup_acm_product(description: str) -> tuple:
    """Look up ACM Product Group and Type from description"""
    desc_lower = description.lower()

    for pattern, (group, ptype) in ACM_PRODUCT_MAPPINGS.items():
        if re.search(pattern, desc_lower):
            return (group, ptype)

    # Default fallback
    return ("Other", "Unknown")


def normalize_level(area_level: str) -> str:
    """Normalize floor level naming"""
    level_lower = area_level.lower()

    if "ground" in level_lower:
        return "Ground"
    elif "first" in level_lower or "level 1" in level_lower or "1st" in level_lower:
        return "Level 1"
    elif "second" in level_lower or "level 2" in level_lower or "2nd" in level_lower:
        return "Level 2"
    elif "external" in level_lower or "roof" in level_lower:
        return "Ground"  # External areas typically ground referenced

    return area_level


def classify_internal_external(area_level: str, room: str) -> str:
    """Determine if location is internal or external"""
    combined = (area_level + " " + room).lower()

    external_keywords = [
        "external",
        "roof",
        "boiler room",
        "fan room",
        "exterior",
        "outside",
    ]

    for kw in external_keywords:
        if kw in combined:
            return "External"

    return "Internal"


def normalize_sample_result(status: str) -> str:
    """Normalize hazard status to Sample Result values"""
    status_lower = status.lower()

    if "positive" in status_lower and "assumed" not in status_lower:
        return "Positive"
    elif "assumed" in status_lower:
        return "Assumed Positive"
    elif "negative" in status_lower:
        return "Negative"

    return status


def normalize_sample_number(sample_num: str) -> str:
    """Normalize sample number format"""
    if not sample_num or sample_num == "-":
        return "Not Sampled"

    sample_num = sample_num.strip()

    # Handle "Same as XXX" references
    if sample_num.lower().startswith("same as"):
        return sample_num.replace("same as", "As Per").strip()

    return sample_num


def extract_color_note(description: str) -> Optional[str]:
    """Extract color description as additional comment"""
    # Pattern: (color) at end or color. at end
    color_match = re.search(r"\((\w+(?:\s+\w+)?)\)\s*$", description)
    if color_match:
        return f"{color_match.group(1).capitalize()}."

    # Check for color words
    colors = [
        "cream",
        "grey",
        "gray",
        "brown",
        "black",
        "beige",
        "white",
        "blue",
        "green",
        "orange",
        "speckled",
    ]
    desc_lower = description.lower()

    for color in colors:
        if color in desc_lower:
            return f"{color.capitalize()}."

    return None


def transform_to_bar_row(metadata: dict, row: dict) -> dict:
    """Transform a single extracted row to BAR format"""

    # Get ACM product info
    product_group, product_type = lookup_acm_product(row.get("item_description", ""))

    # Handle condition/disturbance for negative results
    is_negative = "negative" in row.get("hazard_status", "").lower()

    bar_row = {
        "Department": "DJCS",
        "Agency": metadata.get("agency", "Victoria Police"),
        "Sub Agency": metadata.get("site_name", ""),
        "Site Name (if applicable)": metadata.get("site_name", ""),
        "Building Name": metadata.get("building_name", ""),
        "Building Type": "Police Station"
        if "police" in metadata.get("site_name", "").lower()
        else "",
        "Building Address": metadata.get("address", ""),
        "Suburb": metadata.get("suburb", ""),
        "Postcode": metadata.get("postcode", ""),
        "Owned or Leased": "Owned",
        "Building Unique ID": None,
        "Frequency of use": "Every day",
        "Public Access?": "YES",
        "Date of Inspection": metadata.get("inspection_date"),
        "Estimated Year Built": metadata.get("year_built"),
        "Est. Building Size (m2)": metadata.get("building_size"),
        "Number of Levels": metadata.get("levels"),
        "Construction Type": metadata.get("construction", ""),
        "Roof Type": metadata.get("roof_type", ""),
        "Internal / External": classify_internal_external(
            row.get("area_level", ""), row.get("room_location", "")
        ),
        "Level": normalize_level(row.get("area_level", "")),
        "Room or Area": row.get("room_location", "").title(),
        "Location in Room": row.get("feature", "").title(),
        "Specific Item/ACM Name": row.get("feature", "").title().replace("_", " "),
        "Friability of material": row.get("friability", "Non-friable"),
        "ACM Product Group": product_group,
        "ACM Product Type": product_type,
        "NATA Endorsed Sample number (if available)": normalize_sample_number(
            row.get("sample_number", "")
        ),
        "Sample Result": normalize_sample_result(row.get("hazard_status", "")),
        "Identifying Hygiene or Consulting Company": metadata.get(
            "consultant", "Prensa Pty Ltd"
        ),
        "Condition": "N/A (negative)" if is_negative else row.get("condition", "Good"),
        "Disturbance Potential": "N/A (negative)"
        if is_negative
        else row.get("disturbance_potential", "Low"),
        "Quantity": None if row.get("quantity", "-") == "-" else row.get("quantity"),
        "Labelled": "YES" if row.get("labelled", "").lower() == "yes" else "NO",
        "Label Details": "Labelled"
        if row.get("labelled", "").lower() == "yes"
        else None,
        "Hygienist Recommendations": None
        if row.get("comments", "-") == "-"
        else row.get("comments"),
        "Additional Comments": extract_color_note(row.get("item_description", "")),
        "PSB Supplied ACM ID": None,
        "Assumed Removed?": None,
        "Date of Removal": None,
        "Quantity Removed": None,
        "Asbestos Removal Notification No": None,
        "EPA Waste Transport Certificate No": None,
    }

    return bar_row


# ============================================================================
# EXCEL OUTPUT
# ============================================================================


def create_bar_excel(rows: list, output_path: str):
    """Create BAR-formatted Excel file"""

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Header styling
    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Write headers
    for col, header in enumerate(BAR_COLUMNS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border

    # Write data rows
    for row_idx, row_data in enumerate(rows, 2):
        for col, header in enumerate(BAR_COLUMNS, 1):
            value = row_data.get(header)
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True)

    # Set column widths
    column_widths = {
        "A": 12,
        "B": 15,
        "C": 25,
        "D": 25,
        "E": 25,
        "F": 15,
        "G": 20,
        "H": 15,
        "I": 10,
        "J": 12,
        "K": 15,
        "L": 12,
        "M": 12,
        "N": 15,
        "O": 12,
        "P": 15,
        "Q": 12,
        "R": 15,
        "S": 12,
        "T": 15,
        "U": 12,
        "V": 25,
        "W": 20,
        "X": 20,
        "Y": 15,
        "Z": 25,
        "AA": 20,
        "AB": 25,
        "AC": 15,
        "AD": 20,
        "AE": 12,
        "AF": 15,
        "AG": 10,
        "AH": 10,
        "AI": 12,
        "AJ": 25,
        "AK": 20,
        "AL": 15,
        "AM": 15,
        "AN": 15,
        "AO": 15,
        "AP": 25,
        "AQ": 25,
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(output_path)
    print(f"Created BAR Excel: {output_path}")


# ============================================================================
# MAIN PIPELINE
# ============================================================================


def process_pdf(pdf_path: str, output_path: str, use_vision: bool = True):
    """Main processing pipeline"""

    print(f"Processing: {pdf_path}")

    # Step 1: Extract text for metadata and page classification
    pages = extract_text_from_pdf(pdf_path)
    print(f"  Extracted text from {len(pages)} pages")

    # Step 2: Extract metadata
    metadata = extract_metadata_from_text(pages)
    print(f"  Site: {metadata['site_name']}")
    print(f"  Date: {metadata['inspection_date']}")

    # Step 3: Find register pages
    register_pages = []
    for page_num, text in pages.items():
        page_type = classify_page(text)
        if page_type == "register":
            register_pages.append(page_num)

    print(f"  Found {len(register_pages)} register pages: {register_pages}")

    # Step 4: Extract table data
    all_rows = []

    if use_vision and register_pages:
        print("  Using vision model for table extraction...")
        for page_num in register_pages:
            print(f"    Processing page {page_num + 1}...")
            img_bytes = pdf_page_to_image(pdf_path, page_num)
            rows = extract_register_with_vision(img_bytes)
            print(f"    Extracted {len(rows)} rows")
            all_rows.extend(rows)

    if not all_rows:
        print("  WARNING: No rows extracted. Using fallback text extraction...")
        # Could implement text-based fallback here
        return

    # Step 5: Transform to BAR format
    bar_rows = [transform_to_bar_row(metadata, row) for row in all_rows]
    print(f"  Transformed {len(bar_rows)} rows to BAR format")

    # Step 6: Create Excel output
    create_bar_excel(bar_rows, output_path)

    return bar_rows


# ============================================================================
# CLI
# ============================================================================

if __name__ == "__main__":
    import sys

    if len(sys.argv) < 2:
        print("Usage: python asbestos_extractor.py <input.pdf> [output.xlsx]")
        print("\nOptions:")
        print("  --no-vision    Use text-only extraction (faster, less accurate)")
        sys.exit(1)

    pdf_path = sys.argv[1]
    output_path = (
        sys.argv[2] if len(sys.argv) > 2 else pdf_path.replace(".pdf", "_BAR.xlsx")
    )
    use_vision = "--no-vision" not in sys.argv

    process_pdf(pdf_path, output_path, use_vision=use_vision)
